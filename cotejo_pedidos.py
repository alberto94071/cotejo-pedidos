# ============================================================
#  VERIFICADOR DE PRE-ÓRDENES  v3.3  -  CHRONOS-DEV
#  Autor  : Rony — CHRONOS-DEV  |  rony@chronos-dev.com
#  Versión: 3.3
#  Inst.  : IGSS — Consultorio del Instituto en San Marcos
# ============================================================

import sys, tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, re, os, pandas as pd
from datetime import datetime
from collections import defaultdict
from PIL import Image, ImageTk
import numpy as np
import urllib.request, json, webbrowser

# ── Soporte para recursos empaquetados con PyInstaller (_MEIPASS) ──
def obtener_ruta_recurso(relative_path):
    """Retorna la ruta absoluta del recurso, buscando en _MEIPASS para PyInstaller."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# ── Configurar AppUserModelID para Windows (Agrupación de Barra de Tareas) ──
if os.name == 'nt':
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("chronosdev.verificadorpreordenes.igss.v3.3")
    except Exception:
        pass

# ── Metadatos de la aplicación ─────────────────────────────
APP_VERSION  = '3.9'
APP_AUTHOR   = 'CHRONOS-DEV'
APP_CONTACT  = 'www.chronos-dev.com'
APP_TITLE    = 'Verificador de Pre-Órdenes — CONSULTORIO DEL INSTITUTO EN SAN MARCOS'

# ── Auto-updater — GitHub ───────────────────────────────────
#  Repositorio donde vive version.json y los releases.
#  Cambia el nombre del repo si lo llamás diferente en GitHub.
GITHUB_USER = 'alberto94071'
GITHUB_REPO = 'cotejo-pedidos'
VERSION_URL = (
    f'https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/main/version.json'
)

def check_for_updates():
    """Consulta GitHub en background y retorna (latest_version, download_url) o (None, None)."""
    try:
        req = urllib.request.Request(VERSION_URL, headers={'User-Agent': 'CotejoPedidos'})
        with urllib.request.urlopen(req, timeout=4) as r:
            data = json.loads(r.read().decode('utf-8'))
        latest  = data.get('version', '')
        dl_url  = data.get('download_url', '')
        notes   = data.get('notes', '')
        if latest and latest != APP_VERSION:
            return latest, dl_url, notes
    except Exception:
        pass
    return None, None, None

# ── SubProductos que NUNCA se pueden usar en este tipo de compra ──
SUBPRODUCTOS_PROHIBIDOS = {'001-019-0005', '001-004-0007'}

pdf2image = pytesseract = None

def _load_ocr_libs():
    global pdf2image, pytesseract
    import pdf2image as _p; import pytesseract as _t
    pdf2image = _p; pytesseract = _t
    if os.name == 'nt':
        for ruta in [
            r'C:\Users\elvis.rodriguez\AppData\Local\Programs\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
            os.path.join(os.path.dirname(__file__), 'tesseract', 'tesseract.exe'),
        ]:
            if os.path.exists(ruta):
                pytesseract.pytesseract.tesseract_cmd = ruta; break

# ════════════════════════════════════════════════════════════
#  LOGO IGSS
# ════════════════════════════════════════════════════════════

def _make_igss_logo():
    candidates = [
        obtener_ruta_recurso('igss_logo.png'),
        obtener_ruta_recurso('igss_azul-removebg-preview.png'),
    ]
    path = next((p for p in candidates if os.path.exists(p)), None)
    if path is None: return None
    try:
        img  = Image.open(path).convert('RGBA')
        arr  = np.array(img).astype(float)
        norm = np.clip((arr[:,:,0]+arr[:,:,1]+arr[:,:,2]) / (3*255.0), 0, 1)
        out  = np.zeros_like(arr)
        out[:,:,0] = int(0x7a) * norm
        out[:,:,1] = int(0xff) * norm
        out[:,:,2] = int(0x00) * norm
        out[:,:,3] = (norm * 180).astype(np.uint8)
        return Image.fromarray(out.astype(np.uint8), 'RGBA')
    except Exception:
        return None

# ════════════════════════════════════════════════════════════
#  PARSERS
# ════════════════════════════════════════════════════════════

def _get_poppler_path():
    """Detecta la ruta de Poppler en Windows de forma dinámica."""
    if os.name != 'nt':
        return None
    # 1. Carpeta local junto al script (empaquetado con PyInstaller)
    local = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'poppler', 'bin')
    if os.path.exists(local):
        return local
    # 2. Buscar cualquier versión instalada vía WinGet
    winget_base = os.path.expandvars(
        r'%LOCALAPPDATA%\Microsoft\WinGet\Packages')
    if os.path.isdir(winget_base):
        for entry in os.listdir(winget_base):
            if 'poppler' in entry.lower():
                for sub in ['Library\\bin', 'bin']:
                    candidate = os.path.join(winget_base, entry, sub)
                    if os.path.isdir(candidate):
                        return candidate
                # Buscar un nivel más profundo (ej: poppler-X.Y.Z\Library\bin)
                entry_path = os.path.join(winget_base, entry)
                for inner in os.listdir(entry_path):
                    for sub in ['Library\\bin', 'bin']:
                        candidate = os.path.join(entry_path, inner, sub)
                        if os.path.isdir(candidate):
                            return candidate
    # 3. Rutas comunes
    for p in [
        r'C:\Program Files\poppler\bin',
        r'C:\Program Files (x86)\poppler\bin',
        r'C:\poppler\bin',
        r'C:\tools\poppler\bin',
    ]:
        if os.path.isdir(p):
            return p
    # 4. Dejar que pdf2image use el PATH del sistema
    return None


def parse_pdf(pdf_path, progress_cb=None):
    _load_ocr_libs()
    import os
    os.environ["PATH"] += os.pathsep + r'C:\Users\elvis.rodriguez\AppData\Local\Programs\Tesseract-OCR'
    poppler_kwargs = {}
    p = _get_poppler_path()
    if p:
        poppler_kwargs['poppler_path'] = p
    try:
        pages = pdf2image.convert_from_path(pdf_path, dpi=300, **poppler_kwargs)
    except Exception as e:
        msg = str(e)
        if any(k in msg.lower() for k in ['pdfinfo','poppler','executable']):
            raise RuntimeError("No se encontró Poppler.\nColoca la carpeta 'poppler' junto a este script.\n\n" + msg)
        raise

    total = len(pages); rows = []; correlativo = unidad_id = None
    row_pat  = re.compile(r'(\d{6})\s+.+?\s+(\d{3}-\d{3}-\d{4})\s+([\d,\.]+)\s*$')
    corr_pat = re.compile(r'Correlativo\s+No\.?\s+(\d+/\d+)', re.IGNORECASE)
    uid_pat  = re.compile(r'CENTRO DE COSTO:\s*(\d+)', re.IGNORECASE)

    for idx, page in enumerate(pages, 1):
        if progress_cb: progress_cb(f"Leyendo PDF — página {idx}/{total}…")
        try:
            osd = pytesseract.image_to_osd(page)
            rot = re.search(r'Rotate: (\d+)', osd)
            if rot and int(rot.group(1)) != 0:
                page = page.rotate(-int(rot.group(1)), expand=True)
        except Exception:
            pass
        try:
            text = pytesseract.image_to_string(page, lang='spa', config='--psm 6')
        except Exception as e:
            msg = str(e)
            if any(k in msg.lower() for k in ['tesseract','path','not installed']):
                raise RuntimeError("Tesseract no encontrado.\nhttps://github.com/UB-Mannheim/tesseract/wiki\n\n" + msg)
            raise
        if not correlativo:
            m = corr_pat.search(text)
            if m: correlativo = m.group(1).strip()
        if not unidad_id:
            m = uid_pat.search(text)
            if m: unidad_id = m.group(1).strip()
        for line in text.split('\n'):
            m = row_pat.search(line.strip())
            if m:
                try: rows.append({'codigo': m.group(1), 'subproducto': m.group(2),
                                   'cantidad': float(m.group(3).replace(',','.'))})
                except: pass
    return {'correlativo': correlativo, 'unidad_id': unidad_id, 'rows': rows}


def parse_consolidacion(xls_path):
    with open(xls_path, 'r', encoding='utf-8') as f:
        content = f.read()
    data_vals = re.findall(r'<(?:ss:)?Data[^>]*>([^<]+)</(?:ss:)?Data>', content)
    num = data_vals[1].strip(); headers = data_vals[2:13]
    flat = data_vals[13:]; cols = 11; rows = []; preordenes = set()
    for i in range(0, len(flat), cols):
        row = flat[i:i+cols]
        if len(row) < cols: break
        r = dict(zip(headers, [v.strip() for v in row]))
        if r.get('Pre orden','').startswith('Totales'): break
        try:
            rows.append({'preorden': r.get('Pre orden',''),
                         'cod_insumo': r.get('Cod. Insumo','').strip(),
                         'subproducto': r.get('SubProducto','').strip(),
                         'cant': float(r.get('Cantidad solicitada','0').replace(',','.'))})
            preordenes.add(r.get('Pre orden',''))
        except: pass
    idx_ppr  = {(r['cod_insumo'], r['subproducto']): r['cant'] for r in rows}
    ppr_subs = defaultdict(list)
    for r in rows: ppr_subs[r['cod_insumo']].append(r['subproducto'])
    preorden = next(iter(preordenes),'N/A') if len(preordenes)==1 else ', '.join(preordenes)
    return {'num': num, 'preorden': preorden, 'idx_ppr': idx_ppr, 'ppr_subs': dict(ppr_subs), 'rows': rows}


def extract_year_from_series(series):
    """Extrae de forma robusta el año de 4 dígitos de una serie de fechas de pandas (soporta datetime y strings)."""
    def get_year_val(x):
        if pd.isna(x):
            return None
        if hasattr(x, 'year'):
            return str(x.year)
        s = str(x).strip()
        m = re.search(r'\b(20\d{2})\b', s)
        if m:
            return m.group(1)
        return None
    return series.apply(get_year_val)


def generar_pdf_data_desde_excel(excel_path, uid, corr_num, corr_year):
    df = pd.read_excel(excel_path, header=3)
    df.columns = df.columns.str.strip()
    
    try:
        num_float = float(corr_num)
        filtrado = df[(df['ID Unidad'].astype(str).str.strip() == uid) & (df['Número'] == num_float)].copy()
    except:
        filtrado = df[(df['ID Unidad'].astype(str).str.strip() == uid) & (df['Número'].astype(str).str.strip() == corr_num)].copy()
        
    if corr_year and 'Fecha' in filtrado.columns:
        anios_excel = extract_year_from_series(filtrado['Fecha'])
        filtrado = filtrado[anios_excel == corr_year].copy()
        
    if filtrado.empty:
        raise ValueError(f"No se encontraron registros en el Excel para Unidad ID: {uid}, Correlativo: {corr_num} y Año: {corr_year}.")
        
    rows = []
    for _, r in filtrado.iterrows():
        try:
            rows.append({
                'codigo': str(r['Código Interno']).strip(),
                'subproducto': str(r['Subproductos']).strip(),
                'cantidad': float(r['Cantidad Total'])
            })
        except:
            pass
            
    return {
        'correlativo': f"{corr_num}/{corr_year}",
        'unidad_id': uid,
        'rows': rows
    }


def _find_ppr_col(columns):
    for col in columns:
        if 'ppr' in str(col).lower(): return col
    return None

# ════════════════════════════════════════════════════════════
#  COTEJO TRIPLE
# ════════════════════════════════════════════════════════════

def cotejar_triple(excel_path, pdf_data, consol_data):
    corr_num = pdf_data['correlativo'].split('/')[0] if pdf_data['correlativo'] else None
    corr_year = pdf_data['correlativo'].split('/')[1] if pdf_data['correlativo'] and '/' in pdf_data['correlativo'] else None
    uid      = pdf_data['unidad_id']
    df = pd.read_excel(excel_path, header=3)
    df.columns = df.columns.str.strip()
    df['ID Unidad']      = df['ID Unidad'].astype(str).str.strip()
    df['Número']         = pd.to_numeric(df['Número'], errors='coerce')
    df['Código Interno'] = df['Código Interno'].astype(str).str.strip()
    df['Subproductos']   = df['Subproductos'].astype(str).str.strip()
    if uid is None or corr_num is None:
        raise ValueError("No se pudo extraer Unidad ID o Correlativo del PDF.")

    # Deduplicar las filas del pedido por (codigo, subproducto) para evitar redundancias
    seen_rows = set()
    dedup_rows = []
    for r in pdf_data['rows']:
        row_key = (r['codigo'], r['subproducto'])
        if row_key not in seen_rows:
            dedup_rows.append(r)
            seen_rows.add(row_key)
    pdf_data['rows'] = dedup_rows

    filtrado = df[(df['ID Unidad']==uid) & (df['Número']==float(corr_num))].copy()
    if corr_year and 'Fecha' in filtrado.columns:
        anios_excel = extract_year_from_series(filtrado['Fecha'])
        filtrado = filtrado[anios_excel == corr_year].copy()
        
    if filtrado.empty:
        raise ValueError(f"No hay filas en Excel para\nUnidad: {uid}  |  Correlativo: {pdf_data['correlativo']}")

    ppr_col   = _find_ppr_col(df.columns.tolist())
    tiene_ppr = ppr_col is not None
    ppr_warning = None if tiene_ppr else (
        "⚠️  El Excel NO tiene columna 'Código PPR'.\n"
        "La comparación con Pre orden no será precisa.\n"
        "Usa el reporte completo del sistema.")
    if tiene_ppr: filtrado[ppr_col] = filtrado[ppr_col].astype(str).str.strip()

    excel_idx = {}
    for _, row in filtrado.iterrows():
        key = (str(row['Código Interno']), str(row['Subproductos']))
        ppr = str(row[ppr_col]).strip() if tiene_ppr else None
        excel_idx[key] = {'cant': float(row['Cantidad Total']), 'ppr': ppr}

    idx_ppr  = consol_data['idx_ppr']  if consol_data else {}
    ppr_subs = consol_data['ppr_subs'] if consol_data else {}

    # Pre-calcular la suma de Cantidad PDF agrupada por (PPR, subproducto)
    sum_pdf_by_ppr_sub = defaultdict(float)
    vistos_para_suma = set()  # Para evitar sumas duplicadas de filas con el mismo Código Interno
    for fila in pdf_data['rows']:
        key_excel = (fila['codigo'], fila['subproducto'])
        excel_entry = excel_idx.get(key_excel)
        ppr_code = excel_entry['ppr'] if excel_entry else None
        if ppr_code and ppr_code not in ('None', ''):
            sum_key = (fila['codigo'], fila['subproducto'])
            if sum_key not in vistos_para_suma:
                sum_pdf_by_ppr_sub[(ppr_code, fila['subproducto'])] += fila['cantidad']
                vistos_para_suma.add(sum_key)

    resultados = []; claves_excel_vistas = set()
    sub_incorrectos = []; sub_prohibidos = []

    for fila in pdf_data['rows']:
        key_excel = (fila['codigo'], fila['subproducto'])
        cant_pdf  = fila['cantidad']; claves_excel_vistas.add(key_excel)
        excel_entry = excel_idx.get(key_excel)
        cant_excel  = excel_entry['cant'] if excel_entry else None
        ppr_code    = excel_entry['ppr']  if excel_entry else None
        cant_consol = None; sub_en_consol = None; estado = None

        # ── Verificar subproducto prohibido primero ──
        if fila['subproducto'] in SUBPRODUCTOS_PROHIBIDOS:
            estado = 'SUB_PROHIBIDO'
            sub_prohibidos.append({
                'origen':      'PDF / Pedido',
                'codigo':      fila['codigo'],
                'ppr':         ppr_code or '—',
                'subproducto': fila['subproducto'],
                'cant_pdf':    cant_pdf,
            })

        if estado is None and consol_data and ppr_code and ppr_code not in ('None',''):
            cant_consol = idx_ppr.get((ppr_code, fila['subproducto']))
            if cant_consol is None:
                subs = ppr_subs.get(ppr_code, [])
                if subs:
                    sub_en_consol = ', '.join(subs)
                    # Verificar si alguno de los subproductos en Consolidación está prohibido
                    proh_en_consol = [s for s in subs if s in SUBPRODUCTOS_PROHIBIDOS]
                    if proh_en_consol:
                        estado = 'SUB_PROHIBIDO'
                        for sp_proh in proh_en_consol:
                            sub_prohibidos.append({
                                'origen':      'Pre orden',
                                'codigo':      fila['codigo'],
                                'ppr':         ppr_code or '—',
                                'subproducto': sp_proh,
                                'cant_pdf':    cant_pdf,
                            })
                    else:
                        estado = 'SUB_INCORRECTO'
                        sub_incorrectos.append({'codigo': fila['codigo'], 'ppr': ppr_code,
                                                'sub_pedido': fila['subproducto'],
                                                'sub_consol': sub_en_consol, 'cant_pdf': cant_pdf})
                else:
                    # PPR del Excel no existe en ninguna parte de la pre-orden
                    # Posible código PPR incorrecto ingresado en SIGES
                    estado = 'PPR_NO_EN_CONSOL'
            else:
                # Verificar si la cantidad consol tiene subproducto prohibido
                if fila['subproducto'] in SUBPRODUCTOS_PROHIBIDOS and estado != 'SUB_PROHIBIDO':
                    estado = 'SUB_PROHIBIDO'

        if estado is None:
            # Si hay un PPR válido, comparamos con la suma de las cantidades del PDF que comparten mismo PPR y subproducto
            has_ppr = ppr_code and ppr_code not in ('None', '')
            sum_pdf = sum_pdf_by_ppr_sub.get((ppr_code, fila['subproducto']), cant_pdf) if has_ppr else cant_pdf
            
            ok_e = cant_excel is not None and cant_excel == cant_pdf
            ok_c = cant_consol is None or cant_consol == sum_pdf

            if cant_excel is None:
                estado = 'FALTANTE_EXCEL'
            elif not ok_e and not ok_c:
                estado = 'DIFERENCIA_AMBOS'
            elif not ok_e:
                estado = 'DIFERENCIA_EXCEL'
            elif not ok_c:
                estado = 'DIFERENCIA_CONSOL_MAYOR' if cant_consol > sum_pdf else 'DIFERENCIA_CONSOL_MENOR'
            else:
                estado = 'OK'

        resultados.append({'codigo': fila['codigo'], 'ppr': ppr_code or '—',
                           'subproducto': fila['subproducto'], 'cant_pdf': cant_pdf,
                           'cant_excel': cant_excel, 'cant_consol': cant_consol,
                           'sub_en_consol': sub_en_consol, 'estado': estado})

    # Sobrantes Excel — también verificar prohibidos
    for key, entry in excel_idx.items():
        if key not in claves_excel_vistas:
            ppr_code = entry['ppr']
            cant_consol = idx_ppr.get((ppr_code, key[1])) if consol_data and ppr_code and ppr_code not in ('None','') else None
            est = 'SOBRANTE_EXCEL'
            if key[1] in SUBPRODUCTOS_PROHIBIDOS:
                est = 'SUB_PROHIBIDO'
                sub_prohibidos.append({'origen': 'Excel', 'codigo': key[0], 'ppr': ppr_code or '—',
                                       'subproducto': key[1], 'cant_pdf': None})
            resultados.append({'codigo': key[0], 'ppr': ppr_code or '—', 'subproducto': key[1],
                               'cant_pdf': None, 'cant_excel': entry['cant'], 'cant_consol': cant_consol,
                               'sub_en_consol': None, 'estado': est})

    # Verificar prohibidos en Consolidación
    if consol_data:
        for row in consol_data['rows']:
            if row['subproducto'] in SUBPRODUCTOS_PROHIBIDOS:
                # Solo agregar si no fue ya detectado para este mismo PPR e insumo
                ya = any(sp['ppr'] == row['cod_insumo'] and sp['subproducto'] == row['subproducto'] for sp in sub_prohibidos)
                if not ya:
                    sub_prohibidos.append({'origen': 'Pre orden', 'codigo': '—',
                                           'ppr': row['cod_insumo'], 'subproducto': row['subproducto'],
                                           'cant_pdf': row['cant']})

    return resultados, uid, pdf_data['correlativo'], ppr_warning, sub_incorrectos, sub_prohibidos

# ════════════════════════════════════════════════════════════
#  EXPORTAR
# ════════════════════════════════════════════════════════════

ESTADO_LABELS = {
    'OK':                       '✅ Correcto',
    'DIFERENCIA_EXCEL':         '❌ Diferencia vs Excel',
    'DIFERENCIA_CONSOL_MAYOR':  '⚠️ Pre orden MAYOR que pedido',
    'DIFERENCIA_CONSOL_MENOR':  '❌ Pre orden MENOR que pedido',
    'DIFERENCIA_AMBOS':         '❌ Diferencia en ambos',
    'FALTANTE_EXCEL':           '❌ No encontrado en Excel',
    'SOBRANTE_EXCEL':           '⚠️ Sobrante en Excel (no está en PDF)',
    'SUB_INCORRECTO':           '❌ SubProducto incorrecto en Pre orden',
    'SUB_PROHIBIDO':            '🚫 SUBPRODUCTO NO PERMITIDO para este tipo de compra',
    'PPR_NO_EN_CONSOL':         '⚠️ Código PPR no encontrado en Pre-Orden — verificar código en SIGES',
}

def exportar_reporte(resultados, unidad_id, correlativo,
                     consol_data=None, sub_incorrectos=None, sub_prohibidos=None, outpath=None):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    rows_exp = [{'Código Interno': r['codigo'], 'Código PPR': r['ppr'],
                 'Subproducto': r['subproducto'], 'Cantidad PDF': r['cant_pdf'],
                 'Cantidad Excel': r['cant_excel'], 'Cantidad Pre Orden': r['cant_consol'],
                 'Estado': ESTADO_LABELS.get(r['estado'], r['estado'])} for r in resultados]

    df_out  = pd.DataFrame(rows_exp)
    if not outpath:
        ts      = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname   = f"Cotejo_{unidad_id}_{correlativo.replace('/','_')}_{ts}.xlsx"
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        outpath = os.path.join(desktop if os.path.isdir(desktop) else os.path.expanduser('~'), fname)

    with pd.ExcelWriter(outpath, engine='openpyxl') as writer:
        df_out.to_excel(writer, index=False, sheet_name='Cotejo', startrow=5)
        ws = writer.sheets['Cotejo']
        thin   = Side(style='thin'); border = Border(left=thin, right=thin, top=thin, bottom=thin)
        verde  = PatternFill('solid', fgColor='C6EFCE')
        rojo   = PatternFill('solid', fgColor='FFC7CE')
        amari  = PatternFill('solid', fgColor='FFEB9C')
        gris   = PatternFill('solid', fgColor='D9D9D9')
        rojo_f = PatternFill('solid', fgColor='FF0000')   # rojo fuerte = prohibido

        ws['A1'] = 'REPORTE DE VERIFICACIÓN DE PRE-ÓRDENES'; ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Unidad ID: {unidad_id}'
        ws['A3'] = f'Correlativo pedido: {correlativo}'
        if consol_data: ws['A4'] = f'Pre orden: {consol_data["num"]}   |   Pre-Orden: {consol_data["preorden"]}'
        ws['A5'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}   |   v{APP_VERSION}'

        for cell in ws[6]:
            cell.fill=gris; cell.font=Font(bold=True)
            cell.alignment=Alignment(horizontal='center'); cell.border=border

        for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
            val = str(row[6].value or '')
            if   'NO PERMITIDO' in val:    fill = rojo_f
            elif '✅' in val:              fill = verde
            elif 'SubProducto' in val:     fill = rojo
            elif 'MENOR' in val:           fill = rojo
            elif 'MAYOR' in val:           fill = amari
            elif '❌' in val:              fill = rojo
            elif '⚠️' in val:             fill = amari
            else:                           fill = PatternFill()
            for cell in row:
                cell.fill=fill; cell.border=border
                cell.alignment=Alignment(vertical='center', wrap_text=True)

        for col, w in zip('ABCDEFG', [12,12,18,14,14,15,52]):
            ws.column_dimensions[col].width = w

        # Resumen
        total=len(resultados); ok_c=sum(1 for r in resultados if r['estado']=='OK'); err=total-ok_c
        last=ws.max_row+2
        ws.cell(last,1,'RESUMEN GENERAL').font=Font(bold=True,size=12)
        ws.cell(last+1,1,'Total líneas PDF:'); ws.cell(last+1,2,total)
        ws.cell(last+2,1,'Correctas (✅):').font=Font(color='375623',bold=True); ws.cell(last+2,2,ok_c)
        ws.cell(last+3,1,'Con errores:').font=Font(color='9C0006',bold=True); ws.cell(last+3,2,err)

        # ── Sección SubProductos PROHIBIDOS ──
        if sub_prohibidos:
            sec=ws.max_row+3
            t=ws.cell(sec,1,'🚫  SUBPRODUCTOS NO PERMITIDOS PARA ESTE TIPO DE COMPRA  🚫')
            t.font=Font(bold=True,size=12,color='FFFFFF')
            t.fill=PatternFill('solid',fgColor='CC0000')
            ws.merge_cells(start_row=sec,start_column=1,end_row=sec,end_column=6)
            t.alignment=Alignment(horizontal='center')
            n2=ws.cell(sec+1,1,
                f'Los SubProductos {", ".join(SUBPRODUCTOS_PROHIBIDOS)} NO están autorizados '
                f'para este tipo de compra. Deben eliminarse o corregirse en el sistema.')
            n2.font=Font(italic=True,size=10,color='CC0000')
            ws.merge_cells(start_row=sec+1,start_column=1,end_row=sec+1,end_column=6)
            n2.alignment=Alignment(wrap_text=True); ws.row_dimensions[sec+1].height=30
            enc_r2=sec+2
            for ci,enc in enumerate(['Origen','Código Interno','Código PPR','SubProducto PROHIBIDO','Cantidad','Acción requerida'],1):
                c=ws.cell(enc_r2,ci,enc)
                c.fill=PatternFill('solid',fgColor='CC0000'); c.font=Font(bold=True,color='FFFFFF')
                c.alignment=Alignment(horizontal='center'); c.border=border
            for i,sp in enumerate(sub_prohibidos,1):
                fr=enc_r2+i
                for ci,v in enumerate([sp['origen'],sp['codigo'],sp['ppr'],
                                        sp['subproducto'],sp.get('cant_pdf','—'),
                                        'ELIMINAR — SubProducto no autorizado'],1):
                    c=ws.cell(fr,ci,v)
                    c.fill=PatternFill('solid',fgColor='FFCCCC')
                    c.font=Font(color='CC0000',bold=True); c.border=border
                    c.alignment=Alignment(vertical='center')

        # ── Sección SubProductos incorrectos ──
        if sub_incorrectos:
            sec=ws.max_row+3
            t=ws.cell(sec,1,'⚠️  SUBPRODUCTOS INCORRECTOS EN PRE ORDEN  ⚠️')
            t.font=Font(bold=True,size=12,color='9C0006'); t.fill=PatternFill('solid',fgColor='FFD700')
            ws.merge_cells(start_row=sec,start_column=1,end_row=sec,end_column=6)
            t.alignment=Alignment(horizontal='center')
            n=ws.cell(sec+1,1,'Los siguientes códigos tienen SubProducto diferente en la Pre orden respecto al Pedido. Deben ser corregidos por el Centro de Costo.')
            n.font=Font(italic=True,size=10,color='9C0006')
            ws.merge_cells(start_row=sec+1,start_column=1,end_row=sec+1,end_column=6)
            n.alignment=Alignment(wrap_text=True); ws.row_dimensions[sec+1].height=30
            enc_r=sec+2
            for ci,enc in enumerate(['Código Interno','Código PPR','SubProducto en Pedido','SubProducto en Pre orden','Cantidad','Acción requerida'],1):
                c=ws.cell(enc_r,ci,enc); c.fill=PatternFill('solid',fgColor='9C0006')
                c.font=Font(bold=True,color='FFFFFF'); c.alignment=Alignment(horizontal='center'); c.border=border
            for i,si in enumerate(sub_incorrectos,1):
                fr=enc_r+i
                for ci,v in enumerate([si['codigo'],si['ppr'],si['sub_pedido'],si['sub_consol'],si['cant_pdf'],'Verificar y corregir SubProducto en sistema'],1):
                    c=ws.cell(fr,ci,v); c.fill=PatternFill('solid',fgColor='FFD700')
                    c.border=border; c.alignment=Alignment(vertical='center')
    return outpath

# ════════════════════════════════════════════════════════════
#  NOTIFICACIÓN FLOTANTE
# ════════════════════════════════════════════════════════════

def mostrar_toast(parent, mensaje, duracion_ms=3500):
    toast = tk.Toplevel(parent)
    toast.overrideredirect(True)
    toast.attributes('-topmost', True)
    toast.attributes('-alpha', 0.92)
    toast.configure(bg='#1a251a')
    frame = tk.Frame(toast, bg='#1a251a', highlightbackground='#7aff00',
                     highlightthickness=2, padx=24, pady=14)
    frame.pack()
    tk.Label(frame, text='⚡', font=('Segoe UI', 20), bg='#1a251a', fg='#7aff00').pack(side='left', padx=(0,12))
    tk.Label(frame, text=mensaje, font=('Consolas', 11, 'bold'),
             bg='#1a251a', fg='#f0f0e8', wraplength=320).pack(side='left')
    toast.update_idletasks()
    parent.update_idletasks()
    tw = toast.winfo_reqwidth(); th = toast.winfo_reqheight()
    px = parent.winfo_rootx(); py = parent.winfo_rooty()
    ph = parent.winfo_height()
    x = px + 20; y = py + ph - th - 20
    toast.geometry(f'+{x}+{y}')
    def fade_out(alpha=0.92):
        if alpha <= 0.05: toast.destroy(); return
        toast.attributes('-alpha', alpha)
        toast.after(40, fade_out, alpha - 0.06)
    toast.after(duracion_ms, fade_out)
    return toast


def mostrar_popup_errores(parent, titulo, es_error, subtitulo, cabeceras, filas, pie_pagina):
    """Muestra un cuadro de diálogo modal premium, translúcido (glassmorphism) y sin barra de título del OS."""
    dialog = tk.Toplevel(parent)
    dialog.overrideredirect(True)
    dialog.attributes('-topmost', True)
    dialog.grab_set()
    dialog.focus_set()
    dialog.attributes('-alpha', 0.94)  # Efecto translúcido de glassmorphism
    dialog.configure(bg='#162016')
    
    if es_error == 'gray' or es_error == 'gris':
        border_color = '#888a80'
        icono = '🚫'
    elif es_error is True:
        border_color = '#ff4444'
        icono = '🚫'
    else:
        border_color = '#ffd700'
        icono = '⚠️'
    
    # Limpiar cualquier símbolo duplicado al inicio del título
    clean_title = titulo
    for sym in ['🚫', '⚠️']:
        clean_title = clean_title.replace(sym, '').strip()
    
    # ── Barra de Título Personalizada (Drag handle + Botón Cerrar) ──
    title_bar = tk.Frame(dialog, bg='#1a251a', height=32, highlightbackground=border_color, highlightthickness=1)
    title_bar.pack(fill='x', side='top')
    
    title_lbl = tk.Label(title_bar, text=f"  {icono}  {clean_title}", font=('Segoe UI', 9, 'bold'), bg='#1a251a', fg='#f0f0e8')
    title_lbl.pack(side='left', pady=4)
    
    close_btn = tk.Label(title_bar, text='✕', font=('Segoe UI', 11, 'bold'), bg='#1a251a', fg='#888a80', cursor='hand2')
    close_btn.pack(side='right', padx=12, pady=4)
    close_btn.bind("<Enter>", lambda e: close_btn.config(fg='#ff4444'))
    close_btn.bind("<Leave>", lambda e: close_btn.config(fg='#888a80'))
    close_btn.bind("<Button-1>", lambda e: dialog.destroy())
    
    # Soporte para arrastrar la ventana desde la barra de título personalizada
    def start_drag(event):
        dialog._drag_x = event.x
        dialog._drag_y = event.y
    def drag_motion(event):
        dx = event.x - dialog._drag_x
        dy = event.y - dialog._drag_y
        x = dialog.winfo_x() + dx
        y = dialog.winfo_y() + dy
        dialog.geometry(f"+{x}+{y}")
        
    title_bar.bind("<Button-1>", start_drag)
    title_bar.bind("<B1-Motion>", drag_motion)
    title_lbl.bind("<Button-1>", start_drag)
    title_lbl.bind("<B1-Motion>", drag_motion)
    
    # ── Contenido Principal ──
    main_frame = tk.Frame(dialog, bg='#162016', highlightbackground=border_color,
                          highlightthickness=2, padx=20, pady=20)
    main_frame.pack(fill='both', expand=True)
    
    # Subtítulo e Icono principal
    hdr_frame = tk.Frame(main_frame, bg='#162016')
    hdr_frame.pack(fill='x', pady=(0,15))
    
    tk.Label(hdr_frame, text=icono, font=('Segoe UI', 24), bg='#162016', fg=border_color).pack(side='left', padx=(0,15))
    
    sub_lbl = tk.Label(hdr_frame, text=subtitulo, font=('Segoe UI', 11, 'bold'),
                       bg='#162016', fg='#f0f0e8', justify='left', anchor='w')
    sub_lbl.pack(side='left', fill='x', expand=True)
    
    col_widths = [len(c) for c in cabeceras]
    for row in filas:
        for idx, val in enumerate(row):
            col_widths[idx] = max(col_widths[idx], len(str(val)))
            
    header_str = " | ".join(str(cabeceras[i]).ljust(col_widths[i]) for i in range(len(cabeceras)))
    separator = "─" * (sum(col_widths) + 3 * (len(cabeceras) - 1))
    
    text_content = header_str + "\n" + separator + "\n"
    for row in filas:
        row_str = " | ".join(str(row[i]).ljust(col_widths[i]) for i in range(len(row)))
        text_content += row_str + "\n"
        
    txt_frame = tk.Frame(main_frame, bg='#1a251a', highlightbackground='#2a3a2a', highlightthickness=1)
    txt_frame.pack(fill='both', expand=True, pady=10)
    
    scroll_y = tk.Scrollbar(txt_frame, orient='vertical')
    scroll_x = tk.Scrollbar(txt_frame, orient='horizontal')
    
    h_lines = min(max(len(filas) + 3, 6), 16)
    
    txt = tk.Text(txt_frame, font=('Consolas', 10), bg='#1a251a', fg='#7aff00' if es_error else '#ffd700',
                  height=h_lines, width=sum(col_widths) + 6, wrap='none',
                  yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set,
                  relief='flat', padx=10, pady=10, bd=0)
    
    scroll_y.config(command=txt.yview)
    scroll_x.config(command=txt.xview)
    
    txt.insert('1.0', text_content.strip())
    txt.config(state='disabled')
    
    txt.grid(row=0, column=0, sticky='nsew')
    scroll_y.grid(row=0, column=1, sticky='ns')
    scroll_x.grid(row=1, column=0, sticky='ew')
    
    txt_frame.rowconfigure(0, weight=1)
    txt_frame.columnconfigure(0, weight=1)
    
    if pie_pagina:
        tk.Label(main_frame, text=pie_pagina, font=('Segoe UI', 9, 'italic'),
                 bg='#162016', fg='#888a80', anchor='w').pack(fill='x', pady=(5,15))
                 
    btn_frame = tk.Frame(main_frame, bg='#162016')
    btn_frame.pack(fill='x', side='bottom')
    
    if es_error == 'gray' or es_error == 'gris':
        btn_color = '#888a80'
    elif es_error is True:
        btn_color = '#ff4444'
    else:
        btn_color = '#ffd700'
    text_color = '#131a13'
    
    btn = tk.Button(btn_frame, text='Aceptar', font=('Segoe UI', 10, 'bold'),
                    bg=btn_color, fg=text_color, activebackground=border_color, activeforeground=text_color,
                    relief='flat', padx=24, pady=5, cursor='hand2', command=dialog.destroy)
    btn.pack(side='right')
    
    dialog.update_idletasks()
    
    pw = parent.winfo_width(); ph = parent.winfo_height()
    px = parent.winfo_rootx(); py = parent.winfo_rooty()
    
    # Calcular tamaños geométricos óptimos
    dw = max(dialog.winfo_reqwidth(), sum(col_widths)*8 + 60)
    dh = dialog.winfo_reqheight() + 32  # Sumar el alto de la barra de título personalizada
    
    x = px + (pw - dw) // 2
    y = py + (ph - dh) // 2
    dialog.geometry(f'{dw}x{dh}+{x}+{y}')
    
    dialog.bind('<Escape>', lambda e: dialog.destroy())
    dialog.bind('<Return>', lambda e: dialog.destroy())
    
    dialog.wait_window()


# ════════════════════════════════════════════════════════════
# VERIFICADOR SPS-465 — Parsers, cotejo y exportación
# ════════════════════════════════════════════════════════════

_ESTADOS_SPS = {
    'OK':            '✅  OK — Válida',
    'DUPLICADA':     '🔴  DUPLICADA',
    'YA_PAGADA':     '🟠  YA PAGADA',
    'ANULADA':       '⚠️  ANULADA',
    'NO_REGISTRADA': '❓  NO REGISTRADA',
}

def _parse_igss_sps_reporte(path, progress_cb=None):
    """Lee CSV o Excel del sistema IGSS. Retorna dict {num_sps: datos}."""
    if progress_cb: progress_cb('Leyendo reporte IGSS SPS-465…')
    ext = os.path.splitext(path)[1].lower()

    def _try_read(hdr):
        if ext in ('.xlsx', '.xls'):
            return pd.read_excel(path, header=hdr, dtype=str)
        return pd.read_csv(path, header=hdr, encoding='utf-8-sig',
                           low_memory=False, dtype=str)

    df = None
    for hdr in range(15, -1, -1):
        try:
            tmp = _try_read(hdr)
            cols_upper = [str(c).upper().replace('\n', '').replace('\r', '') for c in tmp.columns]
            if any('NUMERO_SPS' in c.replace('-','') or 'SPS465' in c.replace('_','').replace(' ','').replace('-','')
                   for c in cols_upper):
                df = tmp
                break
        except Exception:
            continue

    if df is None:
        raise ValueError('No se encontró la columna NUMERO_SPS465 en el reporte IGSS.')

    col_sps = next(c for c in df.columns
                   if 'NUMERO_SPS' in str(c).upper().replace('\n', '').replace('\r', '').replace('-','')
                   or 'SPS465' in str(c).upper().replace('\n', '').replace('\r', '').replace('_','').replace(' ','').replace('-',''))

    def _v(row, key):
        if key not in df.columns: return ''
        val = str(row.get(key, '')).strip()
        return '' if val.lower() in ('nan', 'none', '') else val

    lookup = {}
    for _, row in df.iterrows():
        raw = _v(row, col_sps)
        if not raw: continue
        try:    num = str(int(float(raw)))
        except: num = raw
        lookup[num] = {
            'num_sps':          num,
            'proveedor':        _v(row, 'PROVEEDOR1') or _v(row, 'PROVEEDOR'),
            'afiliado':         _v(row, 'AFILIADO'),
            'estudio':          _v(row, 'ESTUDIO'),
            'fecha_emision':    _v(row, 'FECHA_EMISION'),
            'fecha_vencimiento':_v(row, 'FECHA_VENCIMIENTO'),
            'fecha_recepcion':  _v(row, 'FECHA_RECEPCION'),
            'fecha_confrontado':_v(row, 'FECHA_CONFRONTADO'),
            'fecha_pagado':     _v(row, 'FECHA_PAGADO'),
            'fecha_anulado':    _v(row, 'FECHA_ANULADO'),
            'nog':              _v(row, 'NOG'),
            'monto':            _v(row, 'MONTO_IVA'),
        }
    return lookup


def _parse_proveedor_excel_sps(path, progress_cb=None):
    """Lee estadística Excel del proveedor. Retorna lista de dicts."""
    if progress_cb: progress_cb('Leyendo estadística Excel del proveedor…')

    for hdr in range(5):
        try:
            df = pd.read_excel(path, header=hdr, dtype=str)
            # Buscar columna con N/465
            col_465 = None
            for c in df.columns:
                n = str(c).upper().replace(' ','').replace('/','').replace('-','')
                if '465' in n or ('N' in n and 'SPS' in n):
                    col_465 = c; break
            # Fallback: primera columna con mayoría numérica
            if col_465 is None:
                fc = df.columns[0]
                sample = df[fc].dropna().head(15)
                if sum(1 for v in sample
                       if str(v).strip().replace('.','').isdigit()) >= 5:
                    col_465 = fc
            if col_465 is None: continue

            col_nombre = next((c for c in df.columns
                               if any(k in str(c).upper()
                                      for k in ['NOMBRE','APELLIDO','PACIENTE'])), None)
            col_estudio = next((c for c in df.columns
                                if any(k in str(c).upper()
                                       for k in ['ESTUDIO','SERVICIO','EXAMEN'])), None)
            col_fecha = next((c for c in df.columns
                              if 'FECHA' in str(c).upper() and '465' in str(c)), None)

            rows = []
            for _, row in df.iterrows():
                val = str(row[col_465]).strip()
                if val.lower() in ('nan','none','') or not val[:2].isdigit():
                    continue
                try:    num = str(int(float(val.replace(',',''))))
                except: continue
                rows.append({
                    'num_sps': num,
                    'nombre':  str(row[col_nombre]).strip() if col_nombre else '',
                    'estudio': str(row[col_estudio]).strip() if col_estudio else '',
                    'fecha':   str(row[col_fecha]).strip()   if col_fecha   else '',
                })
            if rows: return rows
        except Exception:
            continue
    raise ValueError('No se pudo leer la estadística del proveedor desde el Excel.')


def _parse_proveedor_pdf_sps(path, progress_cb=None):
    """OCR de estadística PDF del proveedor. Retorna lista de dicts."""
    _load_ocr_libs()
    import os
    os.environ["PATH"] += os.pathsep + r'C:\Users\elvis.rodriguez\AppData\Local\Programs\Tesseract-OCR'
    try:
        from img2table.document import Image as Img2TableImage
        from img2table.ocr import TesseractOCR
        HAS_IMG2TABLE = True
    except ImportError:
        HAS_IMG2TABLE = False

    if progress_cb: progress_cb('Convirtiendo PDF del proveedor a imágenes…')

    poppler_kw = {}
    p = _get_poppler_path()
    if p:
        poppler_kw['poppler_path'] = p

    try:
        pages = pdf2image.convert_from_path(path, dpi=300, **poppler_kw)
    except Exception as e:
        raise RuntimeError(f'Error al convertir PDF del proveedor: {e}')

    def _extract_from_text(text, rows_list, seen_set):
        import re
        for line in text.splitlines():
            line = line.strip()
            if not line: continue
            
            m = re.search(r'^.*?(\d{3,6})(?:\b|\D)\s*(?:[|/ilI.,_\"\'\\-]*\s*)*(.*)', line)
            if not m: continue
            num, resto = m.group(1), m.group(2).strip()
            if len(num) < 3: continue
            if num in seen_set: continue
            if any(s in resto.upper() for s in ['TOTAL','SUMA','GRAND','MONTO','CALLE','AVENIDA','ZONA']): continue
            
            nombre = estudio = fecha = ''
            nm = re.match(r'^([A-ZÁÉÍÓÚÑa-záéíóúñ\s,\.]+?)(?:\s+(\d{6,13}))?\s+(.*)', resto)
            if not nm:
                nm = re.match(r'^([A-ZÁÉÍÓÚÑa-záéíóúñ\s,\.]+?)\s+(.*)', resto)
                
            if nm:
                nombre = nm.group(1).strip()
                rest_of_line = nm.group(3) if len(nm.groups()) == 3 and nm.group(3) else nm.group(2)
                if rest_of_line:
                    dm = re.search(r'(\d{1,2}[-/]\w{3}[-/]\d{2,4})', rest_of_line)
                    if dm:
                        fecha = dm.group(1)
                        estudio = rest_of_line[:dm.start()].strip()
                    else:
                        estudio = rest_of_line.strip()
            else:
                nombre = resto
                
            seen_set.add(num)
            rows_list.append({'num_sps': num, 'nombre': nombre, 'estudio': estudio, 'fecha': fecha})

    rows, seen = [], set()
    global_rot = None
    for i, img in enumerate(pages):
        if progress_cb: progress_cb(f'OCR página {i+1}/{len(pages)}…')
        try:
            osd = pytesseract.image_to_osd(img)
            rot = re.search(r'Rotate: (\d+)', osd)
            if rot and int(rot.group(1)) != 0:
                global_rot = int(rot.group(1))
        except Exception:
            pass

        if global_rot:
            img = img.rotate(-global_rot, expand=True)

        text_raw = pytesseract.image_to_string(img, lang='spa', config='--psm 6')
        _extract_from_text(text_raw, rows, seen)

        if HAS_IMG2TABLE:
            try:
                import tempfile, os
                with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as tmp:
                    img.save(tmp.name, 'JPEG')
                    tmp_path = tmp.name
                
                doc = Img2TableImage(tmp_path)
                tess_ocr = TesseractOCR(n_threads=1, lang="spa")
                extracted_tables = doc.extract_tables(ocr=tess_ocr, implicit_rows=True, borderless_tables=True, min_confidence=50)
                if extracted_tables:
                    for table in extracted_tables:
                        lines_img2table = []
                        for idx, row in table.df.iterrows():
                            cols = [str(col).strip().replace('\n', ' ') for col in row if str(col).strip() not in ('nan', 'None', '')]
                            if cols:
                                lines_img2table.append(' '.join(cols))
                        _extract_from_text('\n'.join(lines_img2table), rows, seen)
                os.remove(tmp_path)
            except Exception as e:
                print(f"Error img2table: {e}")

    return rows


def _verificar_sps465(igss_dict, proveedor_rows):
    """Cotejo SPS-465: proveedor vs. registro IGSS."""
    resultados = []
    for item in proveedor_rows:
        num  = str(item['num_sps']).strip()
        igss = igss_dict.get(num, {})
        fc   = igss.get('fecha_confrontado', '').strip()
        fp   = igss.get('fecha_pagado', '').strip()
        fa   = igss.get('fecha_anulado', '').strip()

        if not igss:
            estado  = 'NO_REGISTRADA'
            detalle = 'No aparece en el registro del sistema IGSS.'
        elif fc and fc.lower() not in ('nan','none'):
            estado  = 'DUPLICADA'
            detalle = f'Ya confrontada: {fc}.' + (f' Pagada: {fp}.' if fp and fp.lower() not in ('nan','none') else '')
        elif fp and fp.lower() not in ('nan','none'):
            estado  = 'YA_PAGADA'
            detalle = f'Ya pagada el {fp}.'
        elif fa and fa.lower() not in ('nan','none'):
            estado  = 'ANULADA'
            detalle = f'Anulada el {fa}.'
        else:
            estado  = 'OK'
            detalle = ''

        resultados.append({
            'num_sps':    num,
            'nombre':     item.get('nombre','') or igss.get('afiliado',''),
            'estudio':    item.get('estudio','') or igss.get('estudio',''),
            'fecha_sps':  item.get('fecha','')  or igss.get('fecha_emision',''),
            'proveedor':  igss.get('proveedor',''),
            'nog':        igss.get('nog',''),
            'monto':      igss.get('monto',''),
            'fecha_conf': igss.get('fecha_confrontado',''),
            'fecha_pag':  igss.get('fecha_pagado',''),
            'estado':     estado,
            'detalle':    detalle,
        })
    return resultados


def _exportar_sps465(resultados, proveedor_nombre, num_expediente, output_path):
    """Genera: hoja de cotejo completo + carta A4 de devolución."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Colores ────────────────────────────────────────────────
    C_VERDE  = 'C6EFCE'; C_ROJO   = 'FFC7CE'; C_AMBAR  = 'FFEB9C'
    C_GRIS   = 'D9D9D9'; C_NARANJO= 'FCE4D6'; C_ENCAB  = '1F4E79'
    C_TITULO = '002060'; C_BLANCO = 'FFFFFF'

    def _fill(hex_): return PatternFill('solid', fgColor=hex_)
    def _font(bold=False, color='000000', size=10):
        return Font(name='Calibri', bold=bold, color=color, size=size)
    def _border():
        s = Side(style='thin', color='AAAAAA')
        return Border(left=s, right=s, top=s, bottom=s)
    def _center(wrap=False):
        return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    def _left(wrap=True):
        return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

    color_estado = {
        'OK':           C_VERDE,
        'DUPLICADA':    C_ROJO,
        'YA_PAGADA':    C_NARANJO,
        'ANULADA':      C_AMBAR,
        'NO_REGISTRADA':C_AMBAR,
    }

    # ════════════════════════════════════════════════════════════
    # HOJA 1 — Cotejo completo
    # ════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Cotejo SPS-465'
    ws.freeze_panes = 'A4'

    # Encabezado
    ws.merge_cells('A1:J1')
    ws['A1'] = 'COTEJO DE FORMAS SPS-465'
    ws['A1'].font = _font(True, C_BLANCO, 14)
    ws['A1'].fill = _fill(C_TITULO)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:J2')
    meta = f'Proveedor: {proveedor_nombre or "—"}   |   Expediente: {num_expediente or "—"}   |   Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'] = meta
    ws['A2'].font = _font(False, '444444', 9)
    ws['A2'].alignment = _left(False)
    ws.row_dimensions[2].height = 16

    headers = ['N°/465','Afiliado / Paciente','Estudio','Fecha SPS','Proveedor',
               'NOG','Monto IVA','F. Confrontado','F. Pagado','Estado']
    widths  = [10, 32, 40, 14, 32, 12, 12, 20, 20, 22]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(3, ci, h)
        cell.font = _font(True, C_BLANCO, 10)
        cell.fill = _fill(C_ENCAB)
        cell.alignment = _center(True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    for r, item in enumerate(resultados, 4):
        vals = [item['num_sps'], item['nombre'], item['estudio'], item['fecha_sps'],
                item['proveedor'], item['nog'], item['monto'],
                item['fecha_conf'], item['fecha_pag'],
                _ESTADOS_SPS.get(item['estado'], item['estado'])]
        bg = color_estado.get(item['estado'], C_BLANCO)
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.font = _font(size=9)
            cell.fill = _fill(bg)
            cell.alignment = _left(ci in (2,3,5))
            cell.border = _border()
        ws.row_dimensions[r].height = 18

    # Resumen
    row_sum = len(resultados) + 5
    conteos = {e: sum(1 for x in resultados if x['estado']==e)
               for e in _ESTADOS_SPS}
    ws.merge_cells(f'A{row_sum}:J{row_sum}')
    resumen = '  |  '.join(f'{_ESTADOS_SPS[e]}: {conteos[e]}' for e in _ESTADOS_SPS)
    ws[f'A{row_sum}'] = resumen
    ws[f'A{row_sum}'].font = _font(True, '000000', 9)
    ws[f'A{row_sum}'].fill = _fill(C_GRIS)
    ws[f'A{row_sum}'].alignment = _left(True)

    # ════════════════════════════════════════════════════════════
    # HOJA 2 — Carta de devolución A4
    # ════════════════════════════════════════════════════════════
    problemas = [x for x in resultados if x['estado'] != 'OK']

    ws2 = wb.create_sheet('Carta Devolución')
    ws2.page_setup.paperSize  = ws2.PAPERSIZE_A4
    ws2.page_setup.orientation = 'portrait'
    ws2.page_setup.fitToPage  = True
    ws2.page_setup.fitToWidth = 1
    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.page_margins.left  = 0.75
    ws2.page_margins.right = 0.75

    # Ancho de columnas A-G
    for ci, w in enumerate([8,14,32,36,14,18,18], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    row = 1
    def _carta(r, txt, bold=False, size=10, color='000000',
               bg=None, cols='A:G', wrap=True, height=18, align='left'):
        col_end = cols.split(':')[1] if ':' in cols else cols
        ws2.merge_cells(f'A{r}:{col_end}{r}')
        c = ws2[f'A{r}']
        c.value = txt
        c.font  = Font(name='Calibri', bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if bg: c.fill = _fill(bg)
        ws2.row_dimensions[r].height = height
        return r + 1

    row = _carta(row, 'INSTITUTO GUATEMALTECO DE SEGURIDAD SOCIAL',
                 True, 13, C_BLANCO, C_TITULO, height=26, align='center')
    row = _carta(row, 'CONSULTORIO DEL INSTITUTO EN SAN MARCOS',
                 True, 11, C_BLANCO, C_ENCAB, height=22, align='center')
    row = _carta(row, '', height=8)

    fecha_hoy = datetime.now().strftime('%d de %B de %Y').capitalize()
    row = _carta(row, f'San Marcos, {fecha_hoy}', size=10, align='right')
    row = _carta(row, '', height=8)

    row = _carta(row, 'Señores:', bold=True, size=10)
    row = _carta(row, proveedor_nombre or '[NOMBRE DEL PROVEEDOR]', size=10)
    row = _carta(row, 'Ciudad', size=10)
    row = _carta(row, '', height=8)
    row = _carta(row, 'Estimados señores:', bold=True, size=10)
    row = _carta(row, '', height=6)

    cuerpo = (
        f'Por medio de la presente, nos permitimos devolver el expediente '
        f'No. {num_expediente or "___________"}, correspondiente al cobro de '
        f'Formas SPS-465, en virtud de que al realizar la verificación en el '
        f'sistema de registro del IGSS se detectaron las siguientes inconsistencias:'
    )
    row = _carta(row, cuerpo, size=10, height=52, wrap=True)
    row = _carta(row, '', height=8)

    # Tabla de problemas
    th_bg = C_ENCAB
    ths = ['N°/465','Paciente','Estudio','Fecha SPS','Estado','Motivo']
    th_widths = [8,18,34,14,16,20]
    for ci, (h, w) in enumerate(zip(ths, th_widths), 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[row].height = 20
    for ci, h in enumerate(ths, 1):
        ws2.merge_cells(f'{get_column_letter(ci)}{row}:{get_column_letter(ci)}{row}')
        c = ws2.cell(row, ci, h)
        c.font = Font(name='Calibri', bold=True, color=C_BLANCO, size=9)
        c.fill = _fill(th_bg)
        c.alignment = _center(True)
        c.border = _border()
    row += 1

    for item in (problemas or resultados):
        bg = color_estado.get(item['estado'], C_BLANCO)
        vals = [item['num_sps'], item['nombre'][:30], item['estudio'][:45],
                item['fecha_sps'],
                _ESTADOS_SPS.get(item['estado'], item['estado']), item['detalle']]
        ws2.row_dimensions[row].height = 28
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row, ci, str(v))
            c.font = Font(name='Calibri', size=8)
            c.fill = _fill(bg)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border = _border()
        row += 1

    row += 1
    row = _carta(row, '', height=8)
    cierre = (
        'En virtud de lo anterior, se solicita retirar las formas SPS-465 señaladas, '
        'realizar las correcciones pertinentes y presentar el expediente debidamente '
        'corregido para su trámite de pago correspondiente.'
    )
    row = _carta(row, cierre, size=10, height=48, wrap=True)
    row = _carta(row, '', height=8)
    row = _carta(row, 'Atentamente,', size=10)
    row = _carta(row, '', height=40)
    row = _carta(row, '________________________________________________', size=10)
    row = _carta(row, 'ENCARGADO DE COMPRAS Y SUMINISTROS', bold=True, size=10)
    row = _carta(row, 'Consultorio del Instituto en San Marcos', size=10)
    row = _carta(row, 'Instituto Guatemalteco de Seguridad Social', size=10)

    wb.save(output_path)


# ════════════════════════════════════════════════════════════
#  INTERFAZ GRÁFICA
# ════════════════════════════════════════════════════════════

VERDE='#7aff00'; FONDO='#131a13'; SURFACE='#1a251a'
TEXTO='#f0f0e8'; MUTED='#888a80'; ROJO='#ff4444'

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("verificador de pre ordenes")
        self.geometry('1280x760'); self.minsize(1050,620)
        self.configure(bg=FONDO); self.resizable(True,True)
        
        self.update_idletasks()
        if os.name == 'nt':
            try:
                import ctypes
                hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
                style = ctypes.windll.user32.GetWindowLongW(hwnd, -16) # GWL_STYLE = -16
                style &= ~0x00C00000 # WS_CAPTION = 0x00C00000 (Strips title bar)
                ctypes.windll.user32.SetWindowLongW(hwnd, -16, style)
                ctypes.windll.user32.SetWindowPos(hwnd, 0, 0, 0, 0, 0, 0x0027) # SWP_FRAMECHANGED
            except Exception:
                self.overrideredirect(True)
        else:
            self.overrideredirect(True)
            
        self.attributes('-alpha', 0.98)
        self._back_to_menu = False

        # Icono de la ventana
        self.title_ico_img = None
        for ico in ['cotejo_icon.ico', 'igss_logo.png']:
            ico_path = obtener_ruta_recurso(ico)
            if os.path.exists(ico_path):
                try:
                    if ico.endswith('.ico'):
                        self.iconbitmap(ico_path)
                    else:
                        img = ImageTk.PhotoImage(Image.open(ico_path).resize((32,32)))
                        self.iconphoto(True, img)
                    
                    # Cargar logo de 16x16 para la barra de título personalizada
                    pil_img = Image.open(ico_path).convert('RGBA')
                    resized = pil_img.resize((16, 16), Image.LANCZOS)
                    self.title_ico_img = ImageTk.PhotoImage(resized)
                except: pass
                break

        self.excel_path=tk.StringVar(); self.pdf_path=tk.StringVar()
        self.consol_path=tk.StringVar(); self.status_var=tk.StringVar(value='Listo')
        
        # Parámetros manuales de pedido
        self.correlativo_var=tk.StringVar()
        self.anio_var=tk.StringVar(value=str(datetime.now().year))
        
        # Lista de unidades de adscripción — IGSS San Marcos
        self.unidades_list = [
            "120102 — CAJA DEPARTAMENTAL EN SAN MARCOS",
            "120104 — CONSULTORIO DEL INSTITUTO EN SAN MARCOS",
            "120209 — UIA SAN PEDRO SACATEPÉQUEZ, SAN MARCOS",
            "121009 — UIA TEJUTLA, SAN MARCOS",
            "121109 — UIA SAN RAFAEL PIE DE LA CUESTA, SAN MARCOS",
            "122309 — UIA IXCHIGUÁN, SAN MARCOS",
        ]
        self.unidad_var=tk.StringVar(value=self.unidades_list[0])
        
        # Traces para habilitar/deshabilitar COTEJAR de forma reactiva
        self.excel_path.trace_add('write', lambda *args: self._check_ready())
        self.pdf_path.trace_add('write', lambda *args: self._check_ready())
        self.consol_path.trace_add('write', lambda *args: self._check_ready())
        self.correlativo_var.trace_add('write', lambda *args: self._check_ready())
        self.unidad_var.trace_add('write', lambda *args: self._check_ready())

        self.pdf_data=self.consol_data=self.resultados=None
        self._uid=self._corr=self._sub_incorrectos=self._sub_prohibidos=None
        self._logo_img = None
        self._logo_pil = _make_igss_logo()
        self._show_logo = True
        self._build_ui()
        # Verificar actualizaciones en background al iniciar
        threading.Thread(target=self._check_updates_bg, daemon=True).start()

    def minimize_window(self):
        if os.name == 'nt':
            self.iconify()
        else:
            self.overrideredirect(False)
            self.iconify()
            self.bind("<Map>", self.on_map)

    def on_map(self, event=None):
        self.overrideredirect(True)
        self.unbind("<Map>")
    def toggle_maximize(self):
        if getattr(self, '_is_maximized', False):
            self.geometry(self._prev_geometry)
            self._is_maximized = False
            if hasattr(self, 'max_btn'):
                self.max_btn.config(text='▢')
        else:
            self._prev_geometry = self.geometry()
            self._is_maximized = True
            if hasattr(self, 'max_btn'):
                self.max_btn.config(text='❐')
            w = self.winfo_screenwidth()
            h = self.winfo_screenheight() - 40
            self.geometry(f"{w}x{h}+0+0")
    def _check_updates_bg(self):
        """Corre en thread — si hay update muestra toast en el hilo principal."""
        latest, dl_url, notes = check_for_updates()
        if latest:
            self.after(0, self._mostrar_update_toast, latest, dl_url, notes)

    def _mostrar_update_toast(self, latest, dl_url, notes):
        """Toast especial para actualizaciones — persiste hasta que el usuario lo cierra."""
        toast = tk.Toplevel(self)
        toast.overrideredirect(True)
        toast.attributes('-topmost', True)
        toast.attributes('-alpha', 0.96)
        toast.configure(bg='#1a251a')

        frame = tk.Frame(toast, bg='#1a251a', highlightbackground='#7aff00',
                         highlightthickness=2, padx=20, pady=12)
        frame.pack()

        tk.Label(frame, text='🔄', font=('Segoe UI', 18), bg='#1a251a', fg='#7aff00'
                 ).grid(row=0, column=0, rowspan=2, padx=(0,14))
        tk.Label(frame, text=f'Nueva versión disponible: v{latest}',
                 font=('Consolas', 11, 'bold'), bg='#1a251a', fg='#f0f0e8'
                 ).grid(row=0, column=1, columnspan=2, sticky='w')
        if notes:
            tk.Label(frame, text=notes, font=('Consolas', 9), bg='#1a251a', fg='#888a80',
                     wraplength=280).grid(row=1, column=1, columnspan=2, sticky='w')

        btn_frame = tk.Frame(frame, bg='#1a251a'); btn_frame.grid(row=2, column=1, columnspan=2, pady=(10,0), sticky='e')

        def _descargar():
            webbrowser.open(dl_url); toast.destroy()

        tk.Button(btn_frame, text='⬇  Descargar', font=('Consolas', 9, 'bold'),
                  bg='#7aff00', fg='#131a13', relief='flat', padx=14, pady=4,
                  cursor='hand2', command=_descargar).pack(side='left', padx=(0,8))
        tk.Button(btn_frame, text='Ahora no', font=('Consolas', 9),
                  bg='#2a3a2a', fg='#f0f0e8', relief='flat', padx=10, pady=4,
                  cursor='hand2', command=toast.destroy).pack(side='left')

        toast.update_idletasks()
        px = self.winfo_rootx(); py = self.winfo_rooty()
        ph = self.winfo_height()
        tw = toast.winfo_reqwidth(); th = toast.winfo_reqheight()
        toast.geometry(f'+{px+20}+{py+ph-th-20}')

    def _build_ui(self):
        # ── Barra de Título Personalizada para Ventana Principal ──
        self.title_bar = tk.Frame(self, bg='#0e140e', height=30)
        self.title_bar.pack(fill='x', side='top')
        self.title_bar.pack_propagate(False)
        
        # Título e Icono a la izquierda
        self.ico_lbl = None
        if hasattr(self, 'title_ico_img') and self.title_ico_img:
            self.ico_lbl = tk.Label(self.title_bar, image=self.title_ico_img, bg='#0e140e')
            self.ico_lbl.pack(side='left', padx=(10, 4), pady=4)
            title_text = f" {APP_TITLE}"
        else:
            title_text = f"  ⚡  {APP_TITLE}"

        title_lbl = tk.Label(self.title_bar, text=title_text, font=('Consolas', 9, 'bold'), bg='#0e140e', fg='#f0f0e8')
        title_lbl.pack(side='left', pady=4)
        
        # Controles de ventana interactivos a la derecha
        close_btn = tk.Label(self.title_bar, text='✕', font=('Segoe UI', 11, 'bold'), bg='#0e140e', fg='#888a80', cursor='hand2', width=4)
        close_btn.pack(side='right', fill='y')
        close_btn.bind("<Enter>", lambda e: close_btn.config(bg='#ff4444', fg='#ffffff'))
        close_btn.bind("<Leave>", lambda e: close_btn.config(bg='#0e140e', fg='#888a80'))
        close_btn.bind("<Button-1>", lambda e: self.destroy())
        
        self.max_btn = tk.Label(self.title_bar, text='▢', font=('Segoe UI', 11, 'bold'), bg='#0e140e', fg='#888a80', cursor='hand2', width=4)
        self.max_btn.pack(side='right', fill='y')
        self.max_btn.bind("<Enter>", lambda e: self.max_btn.config(bg='#2a3a2a', fg='#ffffff'))
        self.max_btn.bind("<Leave>", lambda e: self.max_btn.config(bg='#0e140e', fg='#888a80'))
        self.max_btn.bind("<Button-1>", lambda e: self.toggle_maximize())
        
        min_btn = tk.Label(self.title_bar, text='—', font=('Segoe UI', 11, 'bold'), bg='#0e140e', fg='#888a80', cursor='hand2', width=4)
        min_btn.pack(side='right', fill='y')
        min_btn.bind("<Enter>", lambda e: min_btn.config(bg='#2a3a2a', fg='#ffffff'))
        min_btn.bind("<Leave>", lambda e: min_btn.config(bg='#0e140e', fg='#888a80'))
        min_btn.bind("<Button-1>", lambda e: self.minimize_window())
        
        # Funcionalidad de arrastre (drag-and-drop) para mover la ventana principal
        def start_drag(event):
            self._drag_x = event.x
            self._drag_y = event.y
        def drag_motion(event):
            if getattr(self, '_is_maximized', False):
                return
            dx = event.x - self._drag_x
            dy = event.y - self._drag_y
            x = self.winfo_x() + dx
            y = self.winfo_y() + dy
            self.geometry(f"+{x}+{y}")
            
        self.title_bar.bind("<Button-1>", start_drag)
        self.title_bar.bind("<B1-Motion>", drag_motion)
        title_lbl.bind("<Button-1>", start_drag)
        title_lbl.bind("<B1-Motion>", drag_motion)
        if self.ico_lbl:
            self.ico_lbl.bind("<Button-1>", start_drag)
            self.ico_lbl.bind("<B1-Motion>", drag_motion)
            
        # Enlaces de doble clic para maximizar
        self.title_bar.bind("<Double-Button-1>", lambda e: self.toggle_maximize())
        title_lbl.bind("<Double-Button-1>", lambda e: self.toggle_maximize())
        if self.ico_lbl:
            self.ico_lbl.bind("<Double-Button-1>", lambda e: self.toggle_maximize())

        # ── Contenido de la Aplicación ──
        hdr=tk.Frame(self,bg=SURFACE,pady=10); hdr.pack(fill='x')
        tk.Label(hdr,text='CONSULTORIO DEL INSTITUTO',font=('Consolas',16,'bold'),bg=SURFACE,fg=TEXTO).pack(side='left',padx=(18,0))
        tk.Label(hdr,text=' EN SAN MARCOS',           font=('Consolas',16,'bold'),bg=SURFACE,fg=VERDE).pack(side='left')
        tk.Label(hdr,text='  |  Cotejo Triple SIAF',  font=('Segoe UI',10),       bg=SURFACE,fg=MUTED).pack(side='left')
        tk.Button(hdr, text='⬅  Menú Principal', font=('Segoe UI', 9),
                  bg=SURFACE, fg=MUTED, relief='flat', cursor='hand2',
                  command=self._volver_menu).pack(side='right', padx=12, pady=6)

        panel=tk.Frame(self,bg=FONDO,pady=10); panel.pack(fill='x',padx=20)
        self._file_row(panel,'📊  Reporte Excel:',     self.excel_path, self._pick_excel, 0)
        self._file_row(panel,'📄  Forma A-01 PDF:',    self.pdf_path,   self._pick_pdf,   1, optional=True)
        self._file_row(panel,'🗂  Consolidación .xls:', self.consol_path,self._pick_consol,2)

        # Controles manuales (Unidad y Correlativo)
        tk.Label(panel, text='🏢  Unidad de Adscripción:', font=('Segoe UI', 10), bg=FONDO, fg=TEXTO, width=28, anchor='w').grid(row=3, column=0, sticky='w', padx=(0,8), pady=3)
        cb_unidad = ttk.Combobox(panel, textvariable=self.unidad_var, values=self.unidades_list, font=('Consolas', 9), state='normal')
        cb_unidad.grid(row=3, column=1, sticky='ew', padx=(0,8))
        cb_unidad.bind('<<ComboboxSelected>>', lambda e: self._check_ready())

        tk.Label(panel, text='🔢  Correlativo / Año:', font=('Segoe UI', 10), bg=FONDO, fg=TEXTO, width=28, anchor='w').grid(row=4, column=0, sticky='w', padx=(0,8), pady=3)
        corr_frame = tk.Frame(panel, bg=FONDO)
        corr_frame.grid(row=4, column=1, sticky='w', padx=(0,8))

        tk.Entry(corr_frame, textvariable=self.correlativo_var, font=('Consolas', 9), bg=SURFACE, fg=VERDE, insertbackground=VERDE, relief='flat', width=18).pack(side='left')
        tk.Label(corr_frame, text='  /  ', font=('Consolas', 10, 'bold'), bg=FONDO, fg=TEXTO).pack(side='left')

        years = [str(y) for y in range(datetime.now().year + 1, datetime.now().year - 5, -1)]
        cb_year = ttk.Combobox(corr_frame, textvariable=self.anio_var, values=years, font=('Consolas', 9), width=8, state='normal')
        cb_year.pack(side='left')
        cb_year.bind('<<ComboboxSelected>>', lambda e: self._check_ready())

        bf=tk.Frame(self,bg=FONDO); bf.pack(pady=(2,8))
        self.btn_cotejar=tk.Button(bf,text='⚡  COTEJAR',font=('Consolas',12,'bold'),
            bg=VERDE,fg=FONDO,relief='flat',padx=28,pady=7,cursor='hand2',
            command=self._start_cotejo,state='disabled'); self.btn_cotejar.pack(side='left',padx=6)
        self.btn_export=tk.Button(bf,text='📥  Exportar Reporte',font=('Consolas',10),
            bg=SURFACE,fg=TEXTO,relief='flat',padx=18,pady=7,cursor='hand2',
            command=self._exportar,state='disabled'); self.btn_export.pack(side='left',padx=6)

        self.info_bar=tk.Frame(self,bg=SURFACE,pady=5); self.info_bar.pack(fill='x',padx=20)
        self.lbl_info=tk.Label(self.info_bar,text='',font=('Consolas',9),bg=SURFACE,fg=MUTED)
        self.lbl_info.pack(side='left',padx=10)

        # ── Tabla ──
        self.tf_container = tf = tk.Frame(self,bg=FONDO)
        tf.pack(fill='both',expand=True,padx=20,pady=(6,0))

        cols   = ('Cód. Interno','PPR','Subproducto','Cant. PDF','Cant. Excel','Cant. Pre Orden','Estado')
        # ── Anchos fijos + stretch=True en Estado para ocupar todo el recuadro sin cortar texto ──
        widths = [100, 80, 130, 88, 100, 120, 480]
        self.tree=ttk.Treeview(tf,columns=cols,show='headings',height=22)
        for col,w in zip(cols,widths):
            self.tree.heading(col,text=col)
            self.tree.column(col, width=w, minwidth=w, stretch=(col == 'Estado'),
                             anchor='center' if w < 200 else 'w')

        if self._logo_pil:
            self._logo_lbl = tk.Label(tf, bg=SURFACE, bd=0)
            self._logo_lbl.place(relx=0.5, rely=0.5, anchor='center')
            tf.bind('<Configure>', self._reposition_logo)

        vsb=ttk.Scrollbar(tf,orient='vertical',  command=self.tree.yview)
        hsb=ttk.Scrollbar(tf,orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0,column=0,sticky='nsew')
        vsb.grid(row=0,column=1,sticky='ns')
        hsb.grid(row=1,column=0,sticky='ew')
        tf.rowconfigure(0,weight=1); tf.columnconfigure(0,weight=1)

        self.tree.tag_configure('ok',           background='#1a3a1a', foreground='#7aff00')
        self.tree.tag_configure('error',        background='#3a1a1a', foreground='#ff6666')
        self.tree.tag_configure('warning',      background='#3a2a00', foreground='#ffd700')
        self.tree.tag_configure('sub_incorr',   background='#3a1a1a', foreground='#ff6666')
        self.tree.tag_configure('consol_mayor', background='#3a2a00', foreground='#ffd700')
        self.tree.tag_configure('consol_menor', background='#3a1a1a', foreground='#ff6666')
        self.tree.tag_configure('prohibido',    background='#5a0000', foreground='#ff0000')
        self.tree.tag_configure('ppr_no_consol',background='#3a2000', foreground='#ff9900')

        sb=tk.Frame(self,bg=SURFACE,pady=5); sb.pack(fill='x',side='bottom')
        tk.Label(sb,textvariable=self.status_var,font=('Consolas',9),bg=SURFACE,fg=VERDE).pack(side='left',padx=12)
        self.progress=ttk.Progressbar(sb,mode='indeterminate',length=160); self.progress.pack(side='right',padx=6)
        lnk=tk.Label(sb,text=f'v{APP_VERSION}  ·  Desarrollado por CHRONOS-DEV  ↗',
                     font=('Consolas',8),bg=SURFACE,fg=MUTED,cursor='hand2')
        lnk.pack(side='right',padx=12)
        lnk.bind('<Enter>', lambda e: lnk.config(fg=VERDE, font=('Consolas',8,'underline')))
        lnk.bind('<Leave>', lambda e: lnk.config(fg=MUTED,  font=('Consolas',8)))
        lnk.bind('<Button-1>', lambda e: webbrowser.open('https://www.chronos-dev.com'))

        s=ttk.Style(self); s.theme_use('clam')
        s.configure('Treeview',background=SURFACE,foreground=TEXTO,fieldbackground=SURFACE,rowheight=26,font=('Consolas',9))
        s.configure('Treeview.Heading',background='#2a3a2a',foreground=VERDE,font=('Consolas',9,'bold'))
        s.map('Treeview',background=[('selected','#2a4a2a')])
        
        # Estilo premium para los Comboboxes del tema oscuro
        s.configure('TCombobox', fieldbackground=SURFACE, background='#2a3a2a', foreground=VERDE, arrowcolor=VERDE)
        s.map('TCombobox', fieldbackground=[('readonly', SURFACE)], foreground=[('readonly', VERDE)])

    def _reposition_logo(self, event=None):
        if event and event.widget != self.tf_container: return
        if hasattr(self, '_show_logo') and not self._show_logo: return
        if hasattr(self, '_logo_lbl') and hasattr(self, '_logo_pil') and self._logo_pil:
            if event: cw, ch = event.width, event.height
            else:
                self.update_idletasks()
                cw = self.tf_container.winfo_width()
                ch = self.tf_container.winfo_height()
            if cw > 20 and ch > 20:
                scale_dim = min(cw, ch) * 0.80
                orig_w, orig_h = self._logo_pil.size
                ratio = scale_dim / max(orig_w, orig_h)
                new_w = int(orig_w * ratio); new_h = int(orig_h * ratio)
                resized = self._logo_pil.resize((new_w, new_h), Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(resized)
                self._logo_lbl.config(image=self._logo_img)
            self._logo_lbl.place(relx=0.005, rely=0.5, anchor='w')

    def _file_row(self,parent,label,var,cmd,row,optional=False):
        color=MUTED if optional else TEXTO; sfx='  (opcional)' if optional else ''
        tk.Label(parent,text=label+sfx,font=('Segoe UI',10),bg=FONDO,fg=color,width=28,anchor='w').grid(row=row,column=0,sticky='w',padx=(0,8),pady=3)
        tk.Entry(parent,textvariable=var,font=('Consolas',9),bg=SURFACE,fg=VERDE,insertbackground=VERDE,relief='flat',width=60).grid(row=row,column=1,sticky='ew',padx=(0,8))
        tk.Button(parent,text='Examinar…',font=('Segoe UI',9),bg='#2a3a2a',fg=TEXTO,relief='flat',cursor='hand2',command=cmd).grid(row=row,column=2,pady=3)
        parent.columnconfigure(1,weight=1)

    def _pick_excel(self):
        p=filedialog.askopenfilename(title='Reporte Excel',filetypes=[('Excel','*.xlsx *.xls'),('Todos','*.*')])
        if p: self.excel_path.set(p); self._check_ready()

    def _pick_pdf(self):
        p=filedialog.askopenfilename(title='Forma A-01 PDF',filetypes=[('PDF','*.pdf'),('Todos','*.*')])
        if p: self.pdf_path.set(p); self._check_ready()

    def _pick_consol(self):
        p=filedialog.askopenfilename(title='Consolidación',filetypes=[('Excel/XML','*.xls *.xlsx'),('Todos','*.*')])
        if p: self.consol_path.set(p); self._check_ready()

    def _check_ready(self):
        has_excel = bool(self.excel_path.get())
        has_consol = bool(self.consol_path.get())
        has_pdf = bool(self.pdf_path.get())
        has_manual = bool(self.correlativo_var.get().strip()) and bool(self.unidad_var.get().strip())
        
        if has_excel and has_consol and (has_pdf or has_manual):
            self.btn_cotejar.config(state='normal')
        else:
            self.btn_cotejar.config(state='disabled')

    def _start_cotejo(self):
        self._show_logo = True; self._reposition_logo()
        self.btn_cotejar.config(state='disabled'); self.btn_export.config(state='disabled')
        self.lbl_info.config(text='')
        for row in self.tree.get_children(): self.tree.delete(row)
        self.progress.start(12)
        mostrar_toast(self, 'Procesando documentos…\nEsto puede tomar unos segundos.', duracion_ms=4000)
        threading.Thread(target=self._run_cotejo,daemon=True).start()

    def _run_cotejo(self):
        try:
            if self.pdf_path.get():
                self._set_status('Leyendo PDF con OCR…')
                pdf_data=parse_pdf(self.pdf_path.get(),progress_cb=self._set_status)
                if pdf_data:
                    corr = pdf_data['correlativo'] or ''
                    uid_str = pdf_data['unidad_id'] or ''
                    if '/' in corr:
                        corr_num, corr_year = corr.split('/', 1)
                    else:
                        corr_num, corr_year = corr, str(datetime.now().year)
                    
                    matched_unit = None
                    for unit in self.unidades_list:
                        if unit.startswith(uid_str):
                            matched_unit = unit
                            break
                    
                    self.after(0, lambda: [
                        self.correlativo_var.set(corr_num),
                        self.anio_var.set(corr_year),
                        self.unidad_var.set(matched_unit or uid_str)
                    ])
            else:
                self._set_status('Iniciando cotejo manual (leyendo desde Excel)…')
                val_unit = self.unidad_var.get().strip()
                m = re.match(r'^(\d+)', val_unit)
                uid = m.group(1) if m else val_unit
                
                corr_num = self.correlativo_var.get().strip()
                corr_year = self.anio_var.get().strip()
                
                pdf_data = generar_pdf_data_desde_excel(self.excel_path.get(), uid, corr_num, corr_year)
            self.pdf_data=pdf_data
            
            self._set_status('Leyendo Consolidación…')
            consol_data=parse_consolidacion(self.consol_path.get())
            self.consol_data=consol_data
            self._set_status('Cotejando los 3 documentos…')
            resultados,uid,corr,ppr_warning,sub_incorrectos,sub_prohibidos=cotejar_triple(
                self.excel_path.get(),pdf_data,consol_data)
            self.resultados=resultados; self._uid=uid; self._corr=corr
            self._sub_incorrectos=sub_incorrectos; self._sub_prohibidos=sub_prohibidos
            self.after(0,self._mostrar_resultados,resultados,uid,corr,consol_data,
                       ppr_warning,sub_incorrectos,sub_prohibidos)
        except Exception as e:
            self.after(0,self._mostrar_error,str(e))

    def _mostrar_resultados(self,resultados,uid,corr,consol_data,
                             ppr_warning,sub_incorrectos,sub_prohibidos):
        if hasattr(self,'_logo_lbl'): self._logo_lbl.place_forget()
        self._show_logo = False
        self.progress.stop()
        self.btn_cotejar.config(state='normal'); self.btn_export.config(state='normal')
        if ppr_warning: messagebox.showwarning('Sin columna Código PPR',ppr_warning)

        ok=sum(1 for r in resultados if r['estado']=='OK')
        total=len(resultados); errores=total-ok
        ct=f"   |   Pre orden: {consol_data['num']}  Pre-Orden: {consol_data['preorden']}" if consol_data else ''
        self.lbl_info.config(
            text=f'Unidad: {uid}   Correlativo: {corr}{ct}   |   Total: {total}   ✅ OK: {ok}   ❌ Errores: {errores}',
            fg=VERDE if errores==0 else ROJO)

        TAG_MAP={
            'OK':'ok', 'DIFERENCIA_EXCEL':'error', 'DIFERENCIA_AMBOS':'error',
            'FALTANTE_EXCEL':'error', 'SOBRANTE_EXCEL':'warning',
            'SUB_INCORRECTO':'sub_incorr', 'SUB_PROHIBIDO':'prohibido',
            'DIFERENCIA_CONSOL_MAYOR':'consol_mayor', 'DIFERENCIA_CONSOL_MENOR':'consol_menor',
            'PPR_NO_EN_CONSOL':'ppr_no_consol',
        }
        for r in resultados:
            label=ESTADO_LABELS.get(r['estado'],r['estado'])
            if r['estado'] in ('SUB_INCORRECTO', 'SUB_PROHIBIDO') and r.get('sub_en_consol'):
                label+=f"  (Pre orden tiene: {r['sub_en_consol']})"
            self.tree.insert('','end',values=(
                r['codigo'],r['ppr'],r['subproducto'],
                r['cant_pdf']    if r['cant_pdf']    is not None else '—',
                r['cant_excel']  if r['cant_excel']  is not None else '—',
                r['cant_consol'] if r['cant_consol'] is not None else '—',
                label,
            ),tags=(TAG_MAP.get(r['estado'],'ok'),))

        # Popup subproductos prohibidos (prioridad máxima)
        if sub_prohibidos:
            n = len(sub_prohibidos)
            cabeceras = ['Código Interno', 'Código PPR', 'SubProducto PROHIBIDO', 'Origen']
            filas = [[str(sp['codigo']), str(sp['ppr']), str(sp['subproducto']), str(sp['origen'])] for sp in sub_prohibidos]
            
            subt = f"Se detectaron {n} línea(s) con SubProductos que NO están autorizados\npara este tipo de compra. Deben eliminarse o corregirse."
            pie = f"SubProductos prohibidos configurados: {', '.join(SUBPRODUCTOS_PROHIBIDOS)}\nEstos errores se detallan al final del reporte exportado."
            
            mostrar_popup_errores(self, f"🚫 {n} SubProducto(s) NO PERMITIDO(s)", True, subt, cabeceras, filas, pie)

        # Popup subproductos incorrectos
        if sub_incorrectos:
            n = len(sub_incorrectos)
            cabeceras = ['Código Interno', 'Código PPR', 'SubProducto en Pedido', 'SubProducto en Pre orden']
            filas = [[str(si['codigo']), str(si['ppr']), str(si['sub_pedido']), str(si['sub_consol'])] for si in sub_incorrectos]
            
            subt = f"Se detectaron {n} código(s) con un SubProducto diferente en Pre orden\nrespecto al Pedido. Deben ser corregidos en el sistema."
            pie = "Aparecen destacados en rojo y se detallan al final del reporte exportado."
            
            mostrar_popup_errores(self, f"⚠️ {n} SubProducto(s) incorrecto(s)", False, subt, cabeceras, filas, pie)

        msg=(f'✅ Todo correcto ({total} líneas)' if errores==0
             else f'⚠️  {errores} diferencia(s) encontrada(s) de {total}')
        self._set_status(msg)

    def _mostrar_error(self,msg):
        self._show_logo=True; self._reposition_logo()
        self.progress.stop(); self.btn_cotejar.config(state='normal')
        self._set_status(f'Error: {msg}')
        
        # Mostrar el error en la ventana modal premium con tema oscuro
        cabeceras = ['Detalle del error']
        filas = [[str(msg)]]
        subt = "Se ha producido un error durante el procesamiento de la información:"
        pie = "Verifica que el número correlativo, el año, la unidad o los archivos sean válidos."
        mostrar_popup_errores(self, "🚫 Error en la Aplicación", "gray", subt, cabeceras, filas, pie)

    def _exportar(self):
        if not self.resultados: return
        try:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_name = f"Verificacion_PreOrdenes_{self._uid}_{self._corr.replace('/', '_')}_{ts}.xlsx"
            
            path = filedialog.asksaveasfilename(
                title="Guardar Reporte de Verificación",
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
            )
            
            if not path:
                return
                
            exportar_reporte(self.resultados, self._uid, self._corr,
                             self.consol_data, self._sub_incorrectos, self._sub_prohibidos, outpath=path)
            messagebox.showinfo('Reporte guardado', f'Guardado con éxito en:\n{path}')
        except Exception as e:
            messagebox.showerror('Error al exportar', str(e))

    def _volver_menu(self):
        self._back_to_menu = True
        self.destroy()

    def _set_status(self,msg):
        self.after(0,lambda: self.status_var.set(msg))

# ════════════════════════════════════════════════════════════
# VERIFICADOR SPS-465 — Ventana principal
# ════════════════════════════════════════════════════════════

class VerificadorSPS465(tk.Tk):
    def __init__(self):
        super().__init__()
        self._back_to_menu = False
        self.title('Verificador SPS-465')
        self.geometry('1200x740'); self.minsize(980, 600)
        self.configure(bg=FONDO); self.resizable(True, True)
        self.attributes('-alpha', 0.98)

        # ── Icono ────────────────────────────────────────────
        for ico in ['cotejo_icon.ico', 'igss_logo.png']:
            p = obtener_ruta_recurso(ico)
            if os.path.exists(p):
                try:
                    if ico.endswith('.ico'): self.iconbitmap(p)
                    else:
                        img = ImageTk.PhotoImage(Image.open(p).resize((32,32)))
                        self.iconphoto(True, img)
                    break
                except Exception: pass

        # Vars
        self.igss_path    = tk.StringVar()
        self.pdf_path     = tk.StringVar()
        self.excel_prov   = tk.StringVar()
        self.proveedor_var= tk.StringVar()
        self.expediente_var=tk.StringVar()
        self.status_var   = tk.StringVar(value='Listo')

        self._igss_dict   = None
        self._resultados  = None

        for v in (self.igss_path, self.pdf_path, self.excel_prov,
                  self.proveedor_var, self.expediente_var):
            v.trace_add('write', lambda *a: self._check_ready())

        self._build_ui()
        self._check_ready()

    # ── Construcción de UI ───────────────────────────────────
    def _build_ui(self):
        # Título
        bar = tk.Frame(self, bg=SURFACE, height=46)
        bar.pack(fill='x')
        bar.pack_propagate(False)

        logo_img = _make_igss_logo()
        if logo_img:
            try:
                self._logo_tk = ImageTk.PhotoImage(
                    logo_img.resize((32,32), Image.LANCZOS))
                tk.Label(bar, image=self._logo_tk, bg=SURFACE).pack(side='left', padx=(12,6), pady=7)
            except Exception: pass

        tk.Label(bar, text='Verificador SPS-465',
                 font=('Segoe UI', 13, 'bold'), bg=SURFACE, fg=VERDE).pack(side='left', pady=7)
        tk.Label(bar, text='IGSS — San Marcos',
                 font=('Segoe UI', 9), bg=SURFACE, fg=MUTED).pack(side='left', padx=12, pady=7)

        btn_menu = tk.Button(bar, text='⬅  Menú Principal',
                             font=('Segoe UI', 9), bg=SURFACE, fg=MUTED,
                             relief='flat', cursor='hand2',
                             command=self._volver_menu)
        btn_menu.pack(side='right', padx=12, pady=10)

        tk.Frame(self, bg=VERDE, height=2).pack(fill='x')

        # ── Panel de entradas ────────────────────────────────
        panel = tk.Frame(self, bg=FONDO, pady=10, padx=16)
        panel.pack(fill='x')
        panel.columnconfigure(1, weight=1)

        def _fila(parent, label, var, pick_cmd, row, note=''):
            tk.Label(parent, text=label, font=('Segoe UI', 9), bg=FONDO,
                     fg=TEXTO, width=30, anchor='w').grid(row=row, column=0,
                     sticky='w', padx=(0,8), pady=3)
            frm = tk.Frame(parent, bg=FONDO)
            frm.grid(row=row, column=1, sticky='ew', pady=3)
            frm.columnconfigure(0, weight=1)
            ent = tk.Entry(frm, textvariable=var, font=('Consolas',8),
                           bg='#1e2e1e', fg=TEXTO, insertbackground=VERDE,
                           relief='flat', bd=4)
            ent.grid(row=0, column=0, sticky='ew')
            tk.Button(frm, text='📂', font=('Segoe UI',9), bg=SURFACE,
                      fg=VERDE, relief='flat', cursor='hand2',
                      command=pick_cmd).grid(row=0, column=1, padx=(4,0))
            if note:
                tk.Label(parent, text=note, font=('Segoe UI',7),
                         bg=FONDO, fg=MUTED).grid(row=row, column=2,
                         sticky='w', padx=(6,0))

        _fila(panel, '📋  Reporte IGSS (CSV o Excel):', self.igss_path,
              self._pick_igss, 0, note='Requerido')
        _fila(panel, '📄  Estadística Proveedor (PDF / OCR):',
              self.pdf_path, self._pick_pdf, 1, note='Opcional')
        _fila(panel, '📊  Estadística Proveedor (Excel):',
              self.excel_prov, self._pick_excel_prov, 2, note='Opcional')

        # Separador
        tk.Frame(panel, bg=SURFACE, height=1).grid(row=3, column=0,
                 columnspan=3, sticky='ew', pady=(6,4))

        # Proveedor / Expediente
        tk.Label(panel, text='🏢  Nombre del Proveedor:',
                 font=('Segoe UI',9), bg=FONDO, fg=TEXTO,
                 width=30, anchor='w').grid(row=4, column=0, sticky='w', pady=3)
        tk.Entry(panel, textvariable=self.proveedor_var,
                 font=('Segoe UI',9), bg='#1e2e1e', fg=TEXTO,
                 insertbackground=VERDE, relief='flat', bd=4
                 ).grid(row=4, column=1, sticky='ew', pady=3)
        tk.Label(panel, text='Para la carta de devolución',
                 font=('Segoe UI',7), bg=FONDO, fg=MUTED
                 ).grid(row=4, column=2, sticky='w', padx=6)

        tk.Label(panel, text='🗂  Número de Expediente:',
                 font=('Segoe UI',9), bg=FONDO, fg=TEXTO,
                 width=30, anchor='w').grid(row=5, column=0, sticky='w', pady=3)
        tk.Entry(panel, textvariable=self.expediente_var,
                 font=('Segoe UI',9), bg='#1e2e1e', fg=TEXTO,
                 insertbackground=VERDE, relief='flat', bd=4
                 ).grid(row=5, column=1, sticky='ew', pady=3)

        # Botones
        btn_frame = tk.Frame(panel, bg=FONDO)
        btn_frame.grid(row=6, column=0, columnspan=3, pady=(10,0), sticky='e')

        self.btn_verificar = tk.Button(btn_frame, text='🔎  VERIFICAR',
            font=('Segoe UI', 10, 'bold'), bg=VERDE, fg=FONDO,
            relief='flat', padx=18, pady=6, cursor='hand2',
            state='disabled', command=self._run_verificacion)
        self.btn_verificar.pack(side='left', padx=(0,8))

        self.btn_exportar = tk.Button(btn_frame, text='📤  Exportar Reporte',
            font=('Segoe UI', 10), bg=SURFACE, fg=VERDE,
            relief='flat', padx=14, pady=6, cursor='hand2',
            state='disabled', command=self._exportar)
        self.btn_exportar.pack(side='left')

        # ── Treeview ─────────────────────────────────────────
        tree_frame = tk.Frame(self, bg=FONDO)
        tree_frame.pack(fill='both', expand=True, padx=12, pady=(6,0))

        cols = ('num_sps','nombre','estudio','fecha_sps','proveedor','estado','detalle')
        headers = ('N°/465','Paciente / Afiliado','Estudio','Fecha SPS','Proveedor','Estado','Detalle')
        widths  = (80, 200, 260, 100, 180, 150, 260)

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('SPS.Treeview',
            background=SURFACE, foreground=TEXTO,
            fieldbackground=SURFACE, rowheight=24,
            font=('Consolas', 8))
        style.configure('SPS.Treeview.Heading',
            background='#1a3a1a', foreground=VERDE,
            font=('Segoe UI', 9, 'bold'), relief='flat')
        style.map('SPS.Treeview', background=[('selected','#2d5a2d')])

        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings',
                                 style='SPS.Treeview')
        for col, hdr, w in zip(cols, headers, widths):
            self.tree.heading(col, text=hdr)
            self.tree.column(col, width=w, minwidth=60)

        vsb = ttk.Scrollbar(tree_frame, orient='vertical',   command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side='right',  fill='y')
        hsb.pack(side='bottom', fill='x')
        self.tree.pack(fill='both', expand=True)

        # Tags de color
        self.tree.tag_configure('OK',           background='#1a3a1a', foreground='#90ee90')
        self.tree.tag_configure('DUPLICADA',    background='#3a1a1a', foreground='#ff8080')
        self.tree.tag_configure('YA_PAGADA',    background='#3a2a1a', foreground='#ffb870')
        self.tree.tag_configure('ANULADA',      background='#3a3a1a', foreground='#eeee80')
        self.tree.tag_configure('NO_REGISTRADA',background='#2a2a1a', foreground='#dddd60')

        # ── Barra de estado ───────────────────────────────────
        status_bar = tk.Frame(self, bg=SURFACE, height=26)
        status_bar.pack(fill='x', side='bottom')
        status_bar.pack_propagate(False)
        tk.Label(status_bar, textvariable=self.status_var,
                 font=('Segoe UI', 8), bg=SURFACE, fg=MUTED,
                 anchor='w').pack(side='left', padx=10, fill='y')
        self.progress = ttk.Progressbar(status_bar, mode='indeterminate', length=140)
        self.progress.pack(side='right', padx=10, pady=4)

    # ── Selectores de archivo ────────────────────────────────
    def _pick_igss(self):
        p = filedialog.askopenfilename(
            title='Reporte IGSS SPS-465',
            filetypes=[('CSV/Excel','*.csv *.xlsx *.xls'),('Todos','*.*')])
        if p: self.igss_path.set(p)

    def _pick_pdf(self):
        p = filedialog.askopenfilename(
            title='Estadística Proveedor (PDF)',
            filetypes=[('PDF','*.pdf'),('Todos','*.*')])
        if p: self.pdf_path.set(p)

    def _pick_excel_prov(self):
        p = filedialog.askopenfilename(
            title='Estadística Proveedor (Excel)',
            filetypes=[('Excel','*.xlsx *.xls'),('Todos','*.*')])
        if p: self.excel_prov.set(p)

    # ── Validación ───────────────────────────────────────────
    def _check_ready(self):
        has_igss = bool(self.igss_path.get().strip())
        has_prov = bool(self.pdf_path.get().strip() or self.excel_prov.get().strip())
        state = 'normal' if (has_igss and has_prov) else 'disabled'
        if hasattr(self, 'btn_verificar'):
            self.btn_verificar.config(state=state)

    def _set_status(self, msg):
        self.after(0, lambda: self.status_var.set(msg))

    # ── Verificación ─────────────────────────────────────────
    def _run_verificacion(self):
        self.btn_verificar.config(state='disabled')
        self.btn_exportar.config(state='disabled')
        self.tree.delete(*self.tree.get_children())
        self.progress.start(10)

        def _worker():
            try:
                self._set_status('Leyendo reporte IGSS…')
                igss_dict = _parse_igss_sps_reporte(
                    self.igss_path.get(), self._set_status)

                proveedor_rows = []
                if self.pdf_path.get().strip():
                    self._set_status('OCR del PDF del proveedor…')
                    proveedor_rows += _parse_proveedor_pdf_sps(
                        self.pdf_path.get(), self._set_status)
                if self.excel_prov.get().strip():
                    self._set_status('Leyendo Excel del proveedor…')
                    rows_xls = _parse_proveedor_excel_sps(
                        self.excel_prov.get(), self._set_status)
                    # Evitar duplicados si se cargaron ambos archivos
                    nums_ya = {r['num_sps'] for r in proveedor_rows}
                    proveedor_rows += [r for r in rows_xls
                                       if r['num_sps'] not in nums_ya]

                if not proveedor_rows:
                    raise ValueError('No se extrajeron números SPS-465 del proveedor.')

                self._set_status('Cotejando…')
                resultados = _verificar_sps465(igss_dict, proveedor_rows)
                self._igss_dict  = igss_dict
                self._resultados = resultados

                self.after(0, lambda: self._mostrar_resultados(resultados))

            except Exception as exc:
                msg = str(exc)
                self.after(0, lambda m=msg: [
                    messagebox.showerror('Error en verificación', m),
                    self._set_status(f'Error: {m}'),
                    self.btn_verificar.config(state='normal'),
                    self.progress.stop(),
                ])

        threading.Thread(target=_worker, daemon=True).start()

    def _mostrar_resultados(self, resultados):
        self.progress.stop()
        self.tree.delete(*self.tree.get_children())
        cnt = {e: 0 for e in _ESTADOS_SPS}
        for item in resultados:
            estado = item['estado']
            cnt[estado] = cnt.get(estado, 0) + 1
            self.tree.insert('', 'end', values=(
                item['num_sps'], item['nombre'], item['estudio'],
                item['fecha_sps'], item['proveedor'],
                _ESTADOS_SPS.get(estado, estado), item['detalle'],
            ), tags=(estado,))

        resumen = '  |  '.join(
            f'{_ESTADOS_SPS[e]}: {cnt.get(e,0)}' for e in _ESTADOS_SPS)
        self._set_status(resumen)
        self.btn_verificar.config(state='normal')
        self.btn_exportar.config(state='normal')

    # ── Exportar ─────────────────────────────────────────────
    def _exportar(self):
        if not self._resultados:
            messagebox.showwarning('Sin resultados', 'Primero ejecute la verificación.')
            return
        out = filedialog.asksaveasfilename(
            title='Guardar reporte SPS-465',
            defaultextension='.xlsx',
            filetypes=[('Excel','*.xlsx')])
        if not out: return
        try:
            _exportar_sps465(self._resultados,
                              self.proveedor_var.get(),
                              self.expediente_var.get(),
                              out)
            messagebox.showinfo('Exportado', f'Reporte guardado en:\n{out}')
        except Exception as e:
            messagebox.showerror('Error al exportar', str(e))

    # ── Volver al menú ───────────────────────────────────────
    def _volver_menu(self):
        self._back_to_menu = True
        self.destroy()


# ════════════════════════════════════════════════════════════
# LAUNCHER — Menú principal de utilidades IGSS
# Para agregar una nueva herramienta, añade una entrada a TOOLS_REGISTRY
# con su nombre, ícono, descripción y la función que la abre.
# ════════════════════════════════════════════════════════════

TOOLS_REGISTRY = [
    {
        "id": "cotejo",
        "icon": "🔍",
        "name": "Cotejo de\nPre-Órdenes",
        "desc": "Verificación tripartita\nde órdenes de compra",
        "ready": True,
    },
    {
        "id": "sps465",
        "icon": "🏥",
        "name": "Verificador\nSPS-465",
        "desc": "Detección de duplicados\ny faltantes en expedientes",
        "ready": True,
    },
    # ── Agrega nuevas herramientas aquí ──────────────────────
    # {
    #     "id": "mi_nueva_tool",
    #     "icon": "📊",
    #     "name": "Nueva\nHerramienta",
    #     "desc": "Descripción breve\nde la utilidad",
    #     "ready": False,   # False = muestra como "Próximamente"
    # },
]

class LauncherWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self._selected = None
        self.title("IGSS San Marcos — Utilidades")
        self.resizable(False, False)
        self.configure(bg=FONDO)
        self.attributes('-alpha', 0.98)

        # ── Icono ────────────────────────────────────────────
        for ico in ['cotejo_icon.ico', 'igss_logo.png']:
            ico_path = obtener_ruta_recurso(ico)
            if os.path.exists(ico_path):
                try:
                    if ico.endswith('.ico'):
                        self.iconbitmap(ico_path)
                    else:
                        img = ImageTk.PhotoImage(Image.open(ico_path).resize((32, 32)))
                        self.iconphoto(True, img)
                    break
                except Exception:
                    pass

        # ── Centrar en pantalla ──────────────────────────────
        self.update_idletasks()
        cols = min(len(TOOLS_REGISTRY), 4)
        win_w = max(cols * 180 + 60, 460)
        win_h = 360
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{win_w}x{win_h}+{(sw-win_w)//2}+{(sh-win_h)//2}")

        self._build_ui()
        threading.Thread(target=self._check_updates_bg, daemon=True).start()

    def _check_updates_bg(self):
        latest, dl_url, notes = check_for_updates()
        if latest:
            self.after(0, self._mostrar_update_toast, latest, dl_url, notes)

    def _mostrar_update_toast(self, latest, dl_url, notes):
        toast = tk.Toplevel(self)
        toast.overrideredirect(True)
        toast.attributes('-topmost', True)
        toast.attributes('-alpha', 0.96)
        toast.configure(bg='#1a251a')
        frame = tk.Frame(toast, bg='#1a251a', highlightbackground='#7aff00',
                         highlightthickness=2, padx=20, pady=12)
        frame.pack()
        tk.Label(frame, text='🔄', font=('Segoe UI', 18), bg='#1a251a', fg='#7aff00'
                 ).grid(row=0, column=0, rowspan=2, padx=(0, 14))
        tk.Label(frame, text=f'Nueva versión disponible: v{latest}',
                 font=('Consolas', 11, 'bold'), bg='#1a251a', fg='#f0f0e8'
                 ).grid(row=0, column=1, columnspan=2, sticky='w')
        if notes:
            tk.Label(frame, text=notes, font=('Consolas', 9), bg='#1a251a', fg='#888a80',
                     wraplength=280).grid(row=1, column=1, columnspan=2, sticky='w')
        btn_frame = tk.Frame(frame, bg='#1a251a')
        btn_frame.grid(row=2, column=1, columnspan=2, pady=(10, 0), sticky='e')
        tk.Button(btn_frame, text='⬇  Descargar', font=('Consolas', 9, 'bold'),
                  bg='#7aff00', fg='#131a13', relief='flat', padx=14, pady=4,
                  cursor='hand2', command=lambda: (webbrowser.open(dl_url), toast.destroy())
                  ).pack(side='left', padx=(0, 8))
        tk.Button(btn_frame, text='Ahora no', font=('Consolas', 9),
                  bg='#2a3a2a', fg='#f0f0e8', relief='flat', padx=10, pady=4,
                  cursor='hand2', command=toast.destroy).pack(side='left')
        toast.update_idletasks()
        px, py = self.winfo_rootx(), self.winfo_rooty()
        ph = self.winfo_height()
        tw, th = toast.winfo_reqwidth(), toast.winfo_reqheight()
        toast.geometry(f"+{px}+{py + ph - th - 10}")

    # ── UI ───────────────────────────────────────────────────
    def _build_ui(self):
        # Cabecera
        header = tk.Frame(self, bg=SURFACE, pady=0)
        header.pack(fill='x')

        logo_frame = tk.Frame(header, bg=SURFACE)
        logo_frame.pack(side='left', padx=18, pady=12)
        for logo_file in ['igss_logo.png', 'igss_azul-removebg-preview.png']:
            logo_path = obtener_ruta_recurso(logo_file)
            if os.path.exists(logo_path):
                try:
                    img = Image.open(logo_path).resize((48, 48), Image.LANCZOS)
                    self._logo = ImageTk.PhotoImage(img)
                    tk.Label(logo_frame, image=self._logo, bg=SURFACE).pack()
                    break
                except Exception:
                    pass

        title_frame = tk.Frame(header, bg=SURFACE)
        title_frame.pack(side='left', padx=(0, 18), pady=12)
        tk.Label(title_frame, text="IGSS — San Marcos", font=('Segoe UI', 15, 'bold'),
                 bg=SURFACE, fg=VERDE).pack(anchor='w')
        tk.Label(title_frame, text="Centro de Utilidades Administrativas",
                 font=('Segoe UI', 9), bg=SURFACE, fg=MUTED).pack(anchor='w')

        # Separador
        tk.Frame(self, bg=VERDE, height=2).pack(fill='x')

        # Subtítulo
        tk.Label(self, text="Selecciona la herramienta que deseas usar:",
                 font=('Segoe UI', 10), bg=FONDO, fg=TEXTO, pady=14).pack()

        # Grid de herramientas
        grid_frame = tk.Frame(self, bg=FONDO)
        grid_frame.pack(padx=30, pady=(0, 20))

        for i, tool in enumerate(TOOLS_REGISTRY):
            col = i % 4
            row = i // 4
            self._make_card(grid_frame, tool, row, col)

        # Pie
        pie = tk.Frame(self, bg=FONDO); pie.pack(side='bottom', pady=8)
        tk.Label(pie, text=f"v{APP_VERSION}  ·  ", font=('Segoe UI', 8), bg=FONDO, fg=MUTED).pack(side='left')
        lnk = tk.Label(pie, text="Desarrollado por CHRONOS-DEV  ↗", font=('Segoe UI', 8),
                       bg=FONDO, fg=MUTED, cursor='hand2')
        lnk.pack(side='left')
        lnk.bind('<Enter>', lambda e: lnk.config(fg=VERDE, font=('Segoe UI', 8, 'underline')))
        lnk.bind('<Leave>', lambda e: lnk.config(fg=MUTED,  font=('Segoe UI', 8)))
        lnk.bind('<Button-1>', lambda e: webbrowser.open('https://www.chronos-dev.com'))

    def _make_card(self, parent, tool, row, col):
        ready = tool.get("ready", False)
        card_bg = SURFACE if ready else '#111811'
        card_fg = TEXTO if ready else MUTED
        icon_fg = VERDE if ready else '#444'

        card = tk.Frame(parent, bg=card_bg, width=155, height=155,
                        relief='flat', bd=0, cursor='hand2' if ready else 'arrow')
        card.grid(row=row, column=col, padx=8, pady=8)
        card.grid_propagate(False)

        tk.Label(card, text=tool['icon'], font=('Segoe UI', 30),
                 bg=card_bg, fg=icon_fg).place(relx=0.5, rely=0.28, anchor='center')
        tk.Label(card, text=tool['name'], font=('Segoe UI', 10, 'bold'),
                 bg=card_bg, fg=card_fg, justify='center').place(relx=0.5, rely=0.57, anchor='center')
        tk.Label(card, text=tool['desc'] if ready else "Próximamente",
                 font=('Segoe UI', 7), bg=card_bg, fg=MUTED,
                 justify='center').place(relx=0.5, rely=0.80, anchor='center')

        # Borde verde inferior como indicador activo
        if ready:
            tk.Frame(card, bg=VERDE, height=3).place(relx=0, rely=1.0, anchor='sw',
                                                      relwidth=1.0, y=-1)

        if ready:
            def _hover_in(e, c=card):
                for w in [c] + c.winfo_children():
                    try: w.config(bg='#223022')
                    except Exception: pass
            def _hover_out(e, c=card, orig=card_bg):
                for w in [c] + c.winfo_children():
                    try: w.config(bg=orig)
                    except Exception: pass
            for widget in [card] + card.winfo_children():
                widget.bind('<Button-1>', lambda e, tid=tool['id']: self._on_select(tid))
                widget.bind('<Enter>', _hover_in)
                widget.bind('<Leave>', _hover_out)

    def _on_select(self, tool_id):
        self._selected = tool_id
        self.destroy()


# ════════════════════════════════════════════════════════════
def _abrir_herramienta(tool_id):
    """Abre la herramienta indicada. Retorna True si el usuario quiere volver al menú."""
    if tool_id == 'cotejo':
        t = App(); t.mainloop()
        return getattr(t, '_back_to_menu', False)
    if tool_id == 'sps465':
        t = VerificadorSPS465(); t.mainloop()
        return getattr(t, '_back_to_menu', False)
    # elif tool_id == 'mi_nueva_tool':
    #     t = MiNuevaHerramienta(); t.mainloop()
    #     return getattr(t, '_back_to_menu', False)
    return False

if __name__ == '__main__':
    while True:
        launcher = LauncherWindow()
        launcher.mainloop()
        sel = getattr(launcher, '_selected', None)
        if not sel:
            break                      # Cerró el launcher → salir
        volver = _abrir_herramienta(sel)
        if not volver:
            break                      # Cerró la herramienta con X → salir